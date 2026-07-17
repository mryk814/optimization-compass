from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import shutil
import sqlite3
import tempfile
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Literal

from openpyxl import Workbook, load_workbook

from optimization_compass.failure_modes import (
    OBSERVABLE_IDS,
    insert_structured_failure_modes,
)
from optimization_compass.failure_modes import (
    SCENARIOS as FAILURE_SCENARIOS,
)
from optimization_compass.metadata_models import AtlasMetadataSeed
from optimization_compass.predicates import PredicateCatalog
from optimization_compass.problem_instances import ProblemSuiteSeed
from optimization_compass.release_identity import (
    DatasetReleaseIdentity,
    ReleaseIdentityError,
    canonical_identity_json,
    load_dataset_release_identity,
    load_release_authority,
    validate_release_identity,
)
from optimization_compass.versioned_claims import (
    HIGH_USAGE_IMPLEMENTATION_IDS,
    insert_versioned_claims_and_contexts,
)

if TYPE_CHECKING:
    from optimization_compass.release_bundle import ReleaseBundle

DATASET_STEM = "optimization_method_selection_database_v{version}"
COMPACT_RELEASE_ARTIFACT_KEYS = ("ddl", "report", "release_identity")
ROOT = Path(__file__).parents[2]
RELEASE_AUTHORITY_PATH = Path(__file__).parent / "resources/release-authority.json"
RELEASE_AUTHORITY = load_release_authority(RELEASE_AUTHORITY_PATH)
BASE_DATASET_VERSION = RELEASE_AUTHORITY.base_dataset_version
BASE_DATASET_SHA256 = RELEASE_AUTHORITY.base_database_sha256
TARGET_DATASET_VERSION = RELEASE_AUTHORITY.dataset_version
RELEASE_DATE = RELEASE_AUTHORITY.release_date
DEFAULT_MIGRATION = ROOT / "data/migrations/003_atlas_metadata.sql"
DEFAULT_COVERAGE_MIGRATION = ROOT / "data/migrations/004_learning_coverage.sql"
DEFAULT_PREDICATE_MIGRATION = ROOT / "data/migrations/005_atomic_predicates.sql"
DEFAULT_PROBLEM_MIGRATION = ROOT / "data/migrations/006_problem_instances.sql"
DEFAULT_SEMANTIC_VIEW_MIGRATION = ROOT / "data/migrations/007_semantic_view_presets.sql"
DEFAULT_VERSIONED_CLAIMS_MIGRATION = (
    ROOT / "data/migrations/008_versioned_claims_and_benchmark_context.sql"
)
DEFAULT_FAILURE_MODE_MIGRATION = ROOT / "data/migrations/009_structured_failure_modes.sql"
DEFAULT_LEARNING_GRAPH_MIGRATION = ROOT / "data/migrations/010_learning_graph_and_aliases.sql"
DEFAULT_TRF_DEFAULTS_MIGRATION = ROOT / "data/migrations/011_trust_region_reflective_defaults.sql"
DEFAULT_SEED = ROOT / "data/seeds/atlas_metadata.json"
DEFAULT_PREDICATE_SEED = ROOT / "data/seeds/atomic_predicates.json"
DEFAULT_PROBLEM_SEED = ROOT / "src/optimization_compass/resources/problem-suite.json"
LICENSE_BUNDLE = {
    "LICENSE.txt": ROOT / "LICENSE",
    "DATA_LICENSE.txt": ROOT / "DATA_LICENSE",
    "CONTENT_LICENSE.txt": ROOT / "CONTENT_LICENSE",
    "CC-BY-4.0.txt": ROOT / "CC-BY-4.0",
    "NOTICE.txt": ROOT / "NOTICE",
}
DATA_ZIP_LICENSES = ("DATA_LICENSE.txt", "CC-BY-4.0.txt", "NOTICE.txt")
RELEASE_LICENSE_MANIFEST = {
    "code": {
        "spdx_id": "MIT",
        "path": "licenses/LICENSE.txt",
    },
    "data": {
        "spdx_id": "CC-BY-4.0",
        "path": "licenses/DATA_LICENSE.txt",
        "legal_code_path": "licenses/CC-BY-4.0.txt",
    },
    "content": {
        "spdx_id": "CC-BY-4.0",
        "path": "licenses/CONTENT_LICENSE.txt",
        "legal_code_path": "licenses/CC-BY-4.0.txt",
    },
    "notice_path": "licenses/NOTICE.txt",
    "attribution": (
        "Optimization Compass, Copyright 2026 TAKUYA OTANI and Optimization Compass "
        "contributors, https://github.com/mryk814/optimization-compass"
    ),
}
BASE_CHECK_IDS = frozenset(f"CHK{index:03d}" for index in range(1, 13))
ATLAS_CHECK_IDS = frozenset(f"CHK{index:03d}" for index in range(1, 26))
ATLAS_TABLES = frozenset(
    {
        "view_presets",
        "method_visualization_profiles",
        "problem_definitions",
        "problem_definition_archetypes",
        "problem_definition_features",
        "problem_instances",
        "demo_scenarios",
        "comparison_sets",
        "comparison_set_members",
        "learning_edges",
        "terminology_aliases",
        "learning_coverage_expectations",
        "learning_slice_priorities",
        "atomic_predicates",
        "predicate_policies",
        "predicate_coverage",
        "decision_rule_target_retirements",
        "implementation_claims",
        "benchmark_contexts",
        "failure_mode_profiles",
        "failure_mode_triggers",
        "failure_mode_symptoms",
        "failure_mode_diagnostics",
        "failure_mode_mitigations",
        "failure_mode_affected_entities",
        "failure_mode_scenarios",
    }
)


class ReleaseValidationError(ValueError):
    pass


@dataclass(frozen=True)
class LiveCheck:
    check_id: str
    check_name: str
    scope: str
    severity: str
    status: Literal["pass", "warn", "fail", "not_run"]
    observed_value: str
    expected_condition: str
    details: str
    checked_at: str


@dataclass(frozen=True)
class DatabaseVerification:
    ok: bool
    foreign_key_violations: int
    stored_failures: tuple[LiveCheck, ...]
    live_failures: tuple[LiveCheck, ...]
    status_mismatches: tuple[str, ...]
    checks: tuple[LiveCheck, ...]
    dataset_version: str


@dataclass(frozen=True)
class StagedRelease:
    version: str
    output_directory: Path
    database_path: Path
    manifest_path: Path
    release_identity_path: Path
    site_data_directory: Path
    tree_sha256: str


@dataclass(frozen=True)
class FormatVerification:
    ok: bool
    formats: set[str]
    table_count: int


@dataclass(frozen=True)
class Column:
    name: str
    declared_type: str
    pk_order: int


@dataclass(frozen=True)
class TableSnapshot:
    columns: tuple[Column, ...]
    rows: tuple[tuple[Any, ...], ...]


Snapshot = dict[str, TableSnapshot]


def _write_license_bundle(destination: Path) -> None:
    destination.mkdir()
    for name, source in LICENSE_BUNDLE.items():
        if not source.is_file():
            raise ReleaseValidationError(f"required project license is missing: {source.name}")
        shutil.copyfile(source, destination / name)


def _verify_license_bundle(output_directory: Path, manifest: dict[str, Any]) -> None:
    if manifest.get("licenses") != RELEASE_LICENSE_MANIFEST:
        raise ReleaseValidationError("manifest license declarations do not match release policy")
    license_directory = output_directory / "licenses"
    for name, source in LICENSE_BUNDLE.items():
        bundled = license_directory / name
        if not bundled.is_file() or bundled.read_bytes() != source.read_bytes():
            raise ReleaseValidationError(f"release license differs from project notice: {name}")


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def tree_hash(directory: Path) -> str:
    digest = hashlib.sha256()
    for path in sorted(item for item in directory.rglob("*") if item.is_file()):
        relative = path.relative_to(directory).as_posix().encode("utf-8")
        digest.update(len(relative).to_bytes(4, "big"))
        digest.update(relative)
        content = path.read_bytes()
        digest.update(len(content).to_bytes(8, "big"))
        digest.update(content)
    return digest.hexdigest()


def _validate_release_identity(version: str, release_date: str) -> None:
    try:
        validate_release_identity(version, release_date)
    except ReleaseIdentityError as error:
        raise ReleaseValidationError(str(error)) from error


def _validate_output_directory(
    output_directory: Path,
    *,
    protected_inputs: tuple[Path, ...],
) -> None:
    output = output_directory.resolve(strict=False)
    for input_path in protected_inputs:
        protected = input_path.resolve(strict=False)
        if output == protected or output in protected.parents or protected in output.parents:
            raise ReleaseValidationError(
                f"output directory overlaps protected input: {output} and {protected}"
            )
    if output_directory.exists():
        raise ReleaseValidationError(f"output directory must not already exist: {output_directory}")


def build_staged_release(
    base_database: Path,
    output_directory: Path,
    *,
    migration_path: Path = DEFAULT_MIGRATION,
    seed_path: Path = DEFAULT_SEED,
    target_version: str = TARGET_DATASET_VERSION,
    release_date: str = RELEASE_DATE,
) -> StagedRelease:
    _validate_release_identity(target_version, release_date)
    _validate_output_directory(
        output_directory,
        protected_inputs=(
            base_database,
            migration_path,
            DEFAULT_COVERAGE_MIGRATION,
            DEFAULT_PREDICATE_MIGRATION,
            DEFAULT_PROBLEM_MIGRATION,
            DEFAULT_SEMANTIC_VIEW_MIGRATION,
            DEFAULT_VERSIONED_CLAIMS_MIGRATION,
            DEFAULT_FAILURE_MODE_MIGRATION,
            DEFAULT_LEARNING_GRAPH_MIGRATION,
            DEFAULT_TRF_DEFAULTS_MIGRATION,
            seed_path,
            DEFAULT_PREDICATE_SEED,
            DEFAULT_PROBLEM_SEED,
        ),
    )
    _verify_pinned_base(base_database)
    output_directory.mkdir(parents=True)
    stem = DATASET_STEM.format(version=target_version)
    database_path = output_directory / f"{stem}.sqlite"
    shutil.copyfile(base_database, database_path)
    _apply_atlas_metadata(
        database_path,
        migration_path,
        seed_path,
        target_version=target_version,
        release_date=release_date,
    )
    snapshot = read_snapshot(database_path)

    ddl_path = output_directory / f"{stem}_schema.sql"
    json_path = output_directory / f"{stem}.json"
    jsonl_path = output_directory / f"{stem}.jsonl"
    csv_directory = output_directory / f"{stem}_csv"
    csv_zip_path = output_directory / f"{stem}_csv.zip"
    xlsx_path = output_directory / f"{stem}.xlsx"
    report_path = output_directory / f"{stem}_report.md"
    manifest_path = output_directory / f"{stem}_manifest.json"
    release_identity_path = output_directory / f"{stem}_release.json"
    site_data_directory = output_directory / f"{stem}_site-data"

    _write_license_bundle(output_directory / "licenses")
    _write_ddl(database_path, ddl_path)
    _write_json(snapshot, json_path, version=target_version, release_date=release_date)
    _write_jsonl(snapshot, jsonl_path, version=target_version, release_date=release_date)
    _write_csv_directory(snapshot, csv_directory)
    _write_csv_zip(csv_directory, csv_zip_path, release_date=release_date)
    _write_xlsx(snapshot, xlsx_path, release_date=release_date)
    _write_report(
        snapshot,
        report_path,
        version=target_version,
        release_date=release_date,
    )
    from optimization_compass.db import KnowledgeRepository
    from optimization_compass.site_export import export_site_data

    export_site_data(site_data_directory, KnowledgeRepository(database_path))
    release_identity = DatasetReleaseIdentity(
        schema_version=1,
        dataset_version=target_version,
        release_date=release_date,
        database_sha256=sha256_file(database_path),
    )
    release_identity_path.write_text(
        canonical_identity_json(release_identity), encoding="utf-8", newline="\n"
    )
    if release_identity_path.read_bytes() != (site_data_directory / "release.json").read_bytes():
        raise ReleaseValidationError("dataset and site release identities differ")
    manifest = _manifest_payload(
        output_directory,
        stem,
        database_path,
        snapshot,
        version=target_version,
        release_date=release_date,
        include_manifest=False,
    )
    manifest_path.write_text(_canonical_json(manifest, pretty=True), encoding="utf-8")
    verify_release_tree(output_directory)
    return StagedRelease(
        version=target_version,
        output_directory=output_directory,
        database_path=database_path,
        manifest_path=manifest_path,
        release_identity_path=release_identity_path,
        site_data_directory=site_data_directory,
        tree_sha256=tree_hash(output_directory),
    )


def verify_database(path: Path, *, require_atlas: bool = False) -> DatabaseVerification:
    connection = sqlite3.connect(path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    try:
        checked_at = _release_date(connection)
        live = tuple(compute_live_checks(connection, checked_at, require_atlas=require_atlas))
        stored = tuple(_stored_checks(connection))
        foreign_key_violations = len(connection.execute("PRAGMA foreign_key_check").fetchall())
        version = _dataset_version(connection)
        tables = set(_table_names(connection))
    finally:
        connection.close()
    atlas_context = require_atlas or bool(tables & ATLAS_TABLES)
    expected_ids = ATLAS_CHECK_IDS if atlas_context else BASE_CHECK_IDS
    stored_by_id = {check.check_id: check for check in stored}
    live_by_id = {check.check_id: check for check in live}
    mismatches = [
        *(f"missing-stored:{check_id}" for check_id in sorted(expected_ids - stored_by_id.keys())),
        *(f"extra-stored:{check_id}" for check_id in sorted(stored_by_id.keys() - expected_ids)),
        *(f"missing-live:{check_id}" for check_id in sorted(expected_ids - live_by_id.keys())),
        *(f"extra-live:{check_id}" for check_id in sorted(live_by_id.keys() - expected_ids)),
    ]
    mismatches.extend(
        check.check_id
        for check in live
        if check.check_id in expected_ids
        and check.check_id in stored_by_id
        and check.status != stored_by_id[check.check_id].status
    )
    stored_failures = tuple(check for check in stored if check.status == "fail")
    live_failures = tuple(check for check in live if check.status == "fail")
    return DatabaseVerification(
        ok=not foreign_key_violations
        and not stored_failures
        and not live_failures
        and not mismatches,
        foreign_key_violations=foreign_key_violations,
        stored_failures=stored_failures,
        live_failures=live_failures,
        status_mismatches=tuple(mismatches),
        checks=live,
        dataset_version=version,
    )


def compute_live_checks(
    connection: sqlite3.Connection,
    checked_at: str,
    *,
    require_atlas: bool = False,
) -> list[LiveCheck]:
    tables = _table_names(connection)
    total_rows = sum(
        int(connection.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]) for table in tables
    )
    integrity = str(connection.execute("PRAGMA integrity_check").fetchone()[0])
    foreign_keys = connection.execute("PRAGMA foreign_key_check").fetchall()
    duplicate_primary_keys = _duplicate_primary_keys(connection, tables)
    missing_sources = _missing_source_references(connection, tables)
    unresolved_evidence = _unresolved_evidence_targets(connection, tables)
    unresolved_actions = _unresolved_decision_targets(connection)
    blank_foreign_keys = _blank_foreign_keys(connection, tables)
    objective_form_issues = _objective_form_issues(connection)
    release_count, implementation_count = _implementation_release_coverage(connection)
    license_issues = _license_issues(connection)
    maintenance_issues = _maintenance_issues(connection)

    checks = [
        _check(
            "CHK001",
            "SQL DDL creation",
            "all tables",
            "critical",
            integrity == "ok",
            integrity,
            "PRAGMA integrity_check is ok",
            "Schema and b-trees are readable.",
            checked_at,
        ),
        _check(
            "CHK002",
            "SQLite full data load",
            "all base/derived tables",
            "critical",
            total_rows > 0,
            f"{total_rows:,} rows present",
            "all tables are readable and contain released rows",
            "Counted directly from SQLite.",
            checked_at,
        ),
        _check(
            "CHK003",
            "Foreign-key integrity",
            "all declared foreign keys",
            "critical",
            not foreign_keys,
            f"{len(foreign_keys)} violations",
            "PRAGMA foreign_key_check returns zero rows",
            "Computed live; stored status is not trusted.",
            checked_at,
        ),
        _check(
            "CHK004",
            "Primary-key uniqueness",
            "all tables with declared PK",
            "critical",
            not duplicate_primary_keys,
            f"{len(duplicate_primary_keys)} duplicates",
            "zero duplicate primary-key tuples",
            ", ".join(duplicate_primary_keys) or "All declared primary keys are unique.",
            checked_at,
        ),
        _check(
            "CHK005",
            "Source reference integrity",
            "all source reference fields",
            "high",
            not missing_sources,
            f"{len(missing_sources)} unresolved source IDs",
            "every referenced source ID exists",
            ", ".join(missing_sources[:10]) or "Semicolon and JSON source lists included.",
            checked_at,
        ),
        _check(
            "CHK006",
            "Evidence target integrity",
            "evidence_links",
            "high",
            not unresolved_evidence,
            f"{len(unresolved_evidence)} unresolved targets",
            "every target_table/target_id pair resolves",
            ", ".join(unresolved_evidence[:10]) or "Polymorphic evidence targets resolve.",
            checked_at,
        ),
        _check(
            "CHK007",
            "Decision action target typing",
            "decision_rules",
            "high",
            not unresolved_actions,
            f"{len(unresolved_actions)} unresolved IDs",
            "every action target resolves under its declared type",
            ", ".join(unresolved_actions[:10]) or "Decision target interpretation is explicit.",
            checked_at,
        ),
        _check(
            "CHK008",
            "Optional foreign-key null semantics",
            "declared foreign keys",
            "critical",
            not blank_foreign_keys,
            f"{len(blank_foreign_keys)} blank-string foreign keys",
            "absent optional references use SQL NULL",
            ", ".join(blank_foreign_keys[:10]) or "No blank foreign-key values.",
            checked_at,
        ),
        _check(
            "CHK009",
            "Objective-form enum normalization",
            "feature_values + case_feature_map",
            "medium",
            not objective_form_issues,
            f"{len(objective_form_issues)} issues",
            "no free-text discrete objective form remains",
            ", ".join(objective_form_issues) or "Objective-form values are canonical.",
            checked_at,
        ),
        _check(
            "CHK010",
            "Implementation release coverage",
            "implementations",
            "medium",
            True,
            f"{release_count}/{implementation_count} have non-unknown release metadata",
            "unknown remains explicit and tracked",
            "Coverage is advisory and therefore warning when incomplete.",
            checked_at,
            warn=release_count < implementation_count,
        ),
        _check(
            "CHK011",
            "License verification",
            "Manopt MATLAB and NOMAD",
            "high",
            not license_issues,
            f"{len(license_issues)} mismatches",
            "exact license family is recorded",
            ", ".join(license_issues) or "License families match the released assertions.",
            checked_at,
        ),
        _check(
            "CHK012",
            "Maintenance-status risk",
            "JAXopt",
            "high",
            not maintenance_issues,
            f"{len(maintenance_issues)} mismatches",
            "JAXopt legacy status and release are explicit",
            ", ".join(maintenance_issues) or "Maintenance risk is explicit.",
            checked_at,
        ),
    ]
    if not require_atlas and not set(tables) & ATLAS_TABLES:
        return checks

    missing_tables = ATLAS_TABLES - set(tables)
    view_issues = (
        _view_preset_issues(connection) if "view_presets" in tables else ["missing:view_presets"]
    )
    profile_issues = _missing_table_issues(
        tables, {"method_visualization_profiles", "problem_definitions", "problem_instances"}
    )
    if not profile_issues:
        profile_issues = _profile_objective_issues(connection)
    scenario_issues = _missing_table_issues(tables, {"demo_scenarios", "problem_instances"})
    if not scenario_issues:
        scenario_issues = _scenario_issues(connection)
    comparison_issues = _missing_table_issues(
        tables, {"comparison_sets", "comparison_set_members", "problem_instances"}
    )
    if not comparison_issues:
        comparison_issues = _comparison_issues(connection)
    learning_issues = _missing_table_issues(tables, {"learning_edges"})
    if not learning_issues:
        learning_issues = _learning_edge_issues(connection)
    terminology_issues = _missing_table_issues(tables, {"terminology_aliases"})
    if not terminology_issues:
        terminology_issues = _terminology_alias_issues(connection)
    coverage_issues = _missing_table_issues(
        tables, {"learning_coverage_expectations", "learning_slice_priorities"}
    )
    if not coverage_issues:
        coverage_issues = _coverage_expectation_issues(connection)
    explicit_state_issues = _explicit_state_issues(connection, ATLAS_TABLES - missing_tables)
    claim_issues = _versioned_claim_issues(connection, tables)
    benchmark_issues = _benchmark_context_issues(connection, tables)
    failure_mode_issues = _structured_failure_mode_issues(connection, tables)
    checks.extend(
        [
            _check(
                "CHK013",
                "Atlas schema closure",
                f"{len(ATLAS_TABLES)} atlas metadata tables",
                "critical",
                not missing_tables,
                f"{len(missing_tables)} missing tables",
                f"all {len(ATLAS_TABLES)} normalized tables exist",
                ", ".join(sorted(missing_tables)) or "All atlas tables exist.",
                checked_at,
            ),
            _check(
                "CHK014",
                "View preset constraints",
                "view_presets",
                "high",
                not view_issues,
                f"{len(view_issues)} issues",
                "root condition and relation lists are explicit",
                ", ".join(view_issues[:10]) or "Preset roots and relations are valid.",
                checked_at,
            ),
            _check(
                "CHK015",
                "Visualization/problem closure",
                "profiles + problem definitions + instances",
                "critical",
                not profile_issues,
                f"{len(profile_issues)} issues",
                "generators, dimensions, implementations, and JSON are closed",
                ", ".join(profile_issues[:10]) or "Profiles and problems are closed.",
                checked_at,
            ),
            _check(
                "CHK016",
                "Demo scenario closure",
                "demo_scenarios",
                "critical",
                not scenario_issues,
                f"{len(scenario_issues)} issues",
                "scenario references and seed semantics resolve",
                ", ".join(scenario_issues[:10]) or "Scenarios are closed.",
                checked_at,
            ),
            _check(
                "CHK017",
                "Comparison fairness closure",
                "comparison sets and members",
                "critical",
                not comparison_issues,
                f"{len(comparison_issues)} issues",
                "members are normalized and synchronize by oracle evaluations",
                ", ".join(comparison_issues[:10]) or "Comparison membership is normalized.",
                checked_at,
            ),
            _check(
                "CHK018",
                "Learning edge closure",
                "learning_edges",
                "high",
                not learning_issues,
                f"{len(learning_issues)} issues",
                "typed endpoints resolve without self or duplicate edges",
                ", ".join(learning_issues[:10]) or "Learning endpoints resolve.",
                checked_at,
            ),
            _check(
                "CHK019",
                "Explicit state semantics",
                "atlas metadata",
                "critical",
                not explicit_state_issues,
                f"{len(explicit_state_issues)} ambiguous values",
                "unknown, not_applicable, and unsupported are explicit; blanks are forbidden",
                ", ".join(explicit_state_issues[:10]) or "No ambiguous blank state values.",
                checked_at,
            ),
            LiveCheck(
                check_id="CHK020",
                check_name="Release artifact consistency",
                scope="all staged formats",
                severity="critical",
                status="not_run",
                observed_value="database-only verification",
                expected_condition="release tree round-trip and manifest checks pass",
                details="Only verify_release_tree may establish artifact consistency.",
                checked_at=checked_at,
            ),
            _check(
                "CHK021",
                "Learning coverage closure",
                "coverage expectations and priority slices",
                "critical",
                not coverage_issues,
                f"{len(coverage_issues)} issues",
                "subjects, slices, scores, and evidence resolve without inferred status",
                ", ".join(coverage_issues[:10]) or "Coverage policy rows are closed.",
                checked_at,
            ),
            _check(
                "CHK022",
                "Versioned implementation claim closure",
                "implementation_claims",
                "critical",
                not claim_issues,
                f"{len(claim_issues)} issues",
                "every implementation/predicate has one explicit active claim and history resolves",
                ", ".join(claim_issues[:10]) or "Claim ledger and supersession are closed.",
                checked_at,
            ),
            _check(
                "CHK023",
                "Benchmark context completeness",
                "benchmark_contexts + comparison_sets",
                "critical",
                not benchmark_issues,
                f"{len(benchmark_issues)} issues",
                "LP/QP/NLP/MIP/DFO/BO fixtures are complete and comparisons reference context",
                ", ".join(benchmark_issues[:10]) or "Benchmark contexts are complete.",
                checked_at,
            ),
            _check(
                "CHK024",
                "Structured failure-mode closure",
                "failure mode profiles and relations",
                "critical",
                not failure_mode_issues,
                f"{len(failure_mode_issues)} issues",
                "12 profiles resolve triggers, symptoms, diagnostics, mitigations, and scenarios",
                ", ".join(failure_mode_issues[:10]) or "Failure-mode relations are closed.",
                checked_at,
            ),
            _check(
                "CHK025",
                "Terminology alias closure",
                "terminology_aliases",
                "high",
                not terminology_issues,
                f"{len(terminology_issues)} issues",
                "targets and sources resolve; collisions are explicitly disambiguated",
                ", ".join(terminology_issues[:10]) or "Terminology aliases are closed.",
                checked_at,
            ),
        ]
    )
    return checks


def verify_release_tree(output_directory: Path) -> FormatVerification:
    manifests = sorted(output_directory.glob("*_manifest.json"))
    if len(manifests) != 1:
        raise ReleaseValidationError("release tree must contain exactly one manifest")
    manifest, version, release_date = _read_manifest(manifests[0])
    stem = DATASET_STEM.format(version=version)
    if manifests[0].name != f"{stem}_manifest.json":
        raise ReleaseValidationError("manifest filename does not match release identity")
    expected_names = {
        "database": f"{stem}.sqlite",
        "ddl": f"{stem}_schema.sql",
        "json": f"{stem}.json",
        "jsonl": f"{stem}.jsonl",
        "csv_directory": f"{stem}_csv",
        "csv_zip": f"{stem}_csv.zip",
        "xlsx": f"{stem}.xlsx",
        "report": f"{stem}_report.md",
        "release_identity": f"{stem}_release.json",
        "site_data": f"{stem}_site-data",
    }
    if manifest.get("artifacts") != expected_names:
        raise ReleaseValidationError("manifest filenames do not match versioned release contract")
    _verify_license_bundle(output_directory, manifest)
    database_path = output_directory / expected_names["database"]
    reference = read_snapshot(database_path)
    expected_identity = (version, release_date)
    if _read_json_identity(output_directory / expected_names["json"]) != expected_identity:
        raise ReleaseValidationError("json release identity does not match manifest")
    if _read_jsonl_identity(output_directory / expected_names["jsonl"]) != expected_identity:
        raise ReleaseValidationError("jsonl release identity does not match manifest")
    readers: dict[str, Snapshot] = {
        "json": _read_json(output_directory / expected_names["json"]),
        "jsonl": _read_jsonl(output_directory / expected_names["jsonl"]),
        "csv_directory": _read_csv_directory(
            output_directory / expected_names["csv_directory"], reference
        ),
        "csv_zip": _read_csv_zip(output_directory / expected_names["csv_zip"], reference),
        "xlsx": _read_xlsx(output_directory / expected_names["xlsx"], reference),
    }
    for format_name, observed in readers.items():
        if observed != reference:
            raise ReleaseValidationError(f"{format_name} round-trip differs from sqlite")
    _verify_ddl(output_directory / expected_names["ddl"], reference)
    actual_hashes = _artifact_hashes(output_directory, exclude={manifests[0].name})
    if manifest.get("files") != actual_hashes:
        raise ReleaseValidationError("manifest file hashes do not match release tree")
    if manifest.get("table_counts") != {name: len(table.rows) for name, table in reference.items()}:
        raise ReleaseValidationError("manifest table counts do not match sqlite")
    database_result = verify_database(database_path, require_atlas=True)
    if not database_result.ok:
        raise ReleaseValidationError("staged sqlite live release checks failed")
    if _read_database_identity(database_path) != expected_identity:
        raise ReleaseValidationError("sqlite release identity does not match manifest")
    if _read_report_identity(output_directory / expected_names["report"]) != expected_identity:
        raise ReleaseValidationError("report release identity does not match manifest")
    release_identity = load_dataset_release_identity(
        output_directory / expected_names["release_identity"]
    )
    if (release_identity.dataset_version, release_identity.release_date) != expected_identity:
        raise ReleaseValidationError("release identity does not match manifest")
    if release_identity.database_sha256 != sha256_file(database_path):
        raise ReleaseValidationError("release identity database hash does not match sqlite")
    _verify_site_release_tree(
        output_directory / expected_names["site_data"], expected_identity=release_identity
    )
    if manifest.get("database_sha256") != sha256_file(database_path):
        raise ReleaseValidationError("sqlite hash does not match manifest database hash")
    return FormatVerification(
        ok=True,
        formats={"sqlite", "ddl", "json", "jsonl", "csv_directory", "csv_zip", "xlsx"},
        table_count=len(reference),
    )


def _read_manifest(path: Path) -> tuple[dict[str, Any], str, str]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise ReleaseValidationError("manifest is not valid JSON") from error
    if not isinstance(payload, dict):
        raise ReleaseValidationError("manifest schema must be an object")
    expected_types: dict[str, type[Any]] = {
        "schema_version": int,
        "version": str,
        "release_date": str,
        "base_sha256": str,
        "database_sha256": str,
        "artifacts": dict,
        "files": dict,
        "table_counts": dict,
        "validation": dict,
        "licenses": dict,
    }
    for field, expected_type in expected_types.items():
        if type(payload.get(field)) is not expected_type:
            raise ReleaseValidationError(f"manifest schema field is invalid: {field}")
    if payload["schema_version"] != 2:
        raise ReleaseValidationError("unsupported manifest schema version")
    version = payload["version"]
    release_date = payload["release_date"]
    _validate_release_identity(version, release_date)
    for field in ("base_sha256", "database_sha256"):
        if re.fullmatch(r"[0-9a-f]{64}", payload[field]) is None:
            raise ReleaseValidationError(f"manifest schema field is invalid: {field}")
    return payload, version, release_date


def _read_json_identity(path: Path) -> tuple[str, str]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise ReleaseValidationError("json release header is invalid") from error
    if not isinstance(payload, dict):
        raise ReleaseValidationError("json release header is invalid")
    return _validated_payload_identity(payload, format_name="json")


def _read_jsonl_identity(path: Path) -> tuple[str, str]:
    try:
        first_line = path.read_text(encoding="utf-8").splitlines()[0]
        payload = json.loads(first_line)
    except (OSError, IndexError, json.JSONDecodeError) as error:
        raise ReleaseValidationError("jsonl release header is invalid") from error
    if not isinstance(payload, dict) or payload.get("type") != "release":
        raise ReleaseValidationError("jsonl release header is invalid")
    return _validated_payload_identity(payload, format_name="jsonl")


def _validated_payload_identity(payload: dict[str, Any], *, format_name: str) -> tuple[str, str]:
    version = payload.get("version")
    release_date = payload.get("release_date")
    if not isinstance(version, str) or not isinstance(release_date, str):
        raise ReleaseValidationError(f"{format_name} release identity is invalid")
    _validate_release_identity(version, release_date)
    return version, release_date


def _read_database_identity(path: Path) -> tuple[str, str]:
    connection = sqlite3.connect(path)
    try:
        row = connection.execute(
            "SELECT version, release_date FROM version_history "
            "ORDER BY release_date DESC, version DESC LIMIT 1"
        ).fetchone()
    finally:
        connection.close()
    if row is None:
        raise ReleaseValidationError("sqlite release identity is missing")
    version, release_date = str(row[0]), str(row[1])
    _validate_release_identity(version, release_date)
    return version, release_date


def _read_report_identity(path: Path) -> tuple[str, str]:
    lines = path.read_text(encoding="utf-8").splitlines()
    versions = [
        match.group(1) for line in lines if (match := re.fullmatch(r"- Version: `(.+)`", line))
    ]
    release_dates = [
        match.group(1) for line in lines if (match := re.fullmatch(r"- Release date: `(.+)`", line))
    ]
    if len(versions) != 1 or len(release_dates) != 1:
        raise ReleaseValidationError("report release identity is invalid")
    _validate_release_identity(versions[0], release_dates[0])
    return versions[0], release_dates[0]


def _verify_site_release_tree(
    directory: Path, *, expected_identity: DatasetReleaseIdentity
) -> None:
    required_paths = {
        "release.json",
        "manifest.json",
        "content.json",
        "gallery.json",
        "comparisons.json",
        "recommendation/site-data.json",
        "problems.json",
        "traces/index.json",
        "visualization-scenarios.json",
        "search-trees/index.json",
        "coverage.json",
        "learning-graph.json",
    }
    actual_paths = {path.relative_to(directory).as_posix() for path in directory.rglob("*.json")}
    missing = sorted(required_paths - actual_paths)
    if missing:
        raise ReleaseValidationError(f"site release tree is missing required assets: {missing}")
    observed_identity = load_dataset_release_identity(directory / "release.json")
    if observed_identity != expected_identity:
        raise ReleaseValidationError("site release identity does not match dataset release")
    for relative in sorted(actual_paths - {"release.json"}):
        path = directory / relative
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as error:
            raise ReleaseValidationError(f"site asset is not valid JSON: {relative}") from error
        is_embedded_visualization_payload = relative.startswith("visualizations/")
        if not isinstance(payload, dict) or (
            not is_embedded_visualization_payload
            and payload.get("dataset_version") != expected_identity.dataset_version
        ):
            raise ReleaseValidationError(
                f"site asset dataset version does not match release: {relative}"
            )
    manifest = json.loads((directory / "manifest.json").read_text(encoding="utf-8"))
    referenced = {str(item["path"]) for item in manifest["views"] if isinstance(item, dict)}
    referenced.add(str(manifest["recommendation"]["path"]))
    referenced.add(str(manifest["traces"]["path"]))
    referenced.add(str(manifest["problems"]["path"]))
    referenced.add(str(manifest["visualization_scenarios"]["path"]))
    referenced.add(str(manifest["entity_links"]["path"]))
    referenced.add(str(manifest["sources"]["path"]))
    referenced.add(str(manifest["implementation_claims"]["path"]))
    referenced.add(str(manifest["benchmark_contexts"]["path"]))
    referenced.add(str(manifest["failure_modes"]["path"]))
    referenced.add(str(manifest["failure_discovery"]["path"]))
    referenced.add(str(manifest["coverage"]["path"]))
    referenced.add(str(manifest["coverage"]["report_path"]))
    for relative in referenced:
        if not (directory / relative).is_file():
            raise ReleaseValidationError(f"site manifest references missing asset: {relative}")
    trace_index_path = directory / str(manifest["traces"]["path"])
    if manifest["traces"]["bytes"] != trace_index_path.stat().st_size:
        raise ReleaseValidationError("site trace index byte count does not match manifest")
    if manifest["traces"]["sha256"] != sha256_file(trace_index_path):
        raise ReleaseValidationError("site trace index hash does not match manifest")
    scenario_index_path = directory / str(manifest["visualization_scenarios"]["path"])
    scenario_index = json.loads(scenario_index_path.read_text(encoding="utf-8"))
    for scenario in scenario_index["scenarios"]:
        artifact = scenario["artifact"]
        payload_path = directory / str(artifact["payload_path"])
        if not payload_path.is_file():
            raise ReleaseValidationError(
                f"visualization scenario references missing payload: {artifact['payload_path']}"
            )
        if payload_path.stat().st_size != artifact["payload_bytes"]:
            raise ReleaseValidationError(
                f"visualization scenario payload byte count differs: {artifact['payload_path']}"
            )
        if sha256_file(payload_path) != artifact["payload_sha256"]:
            raise ReleaseValidationError(
                f"visualization scenario payload hash differs: {artifact['payload_path']}"
            )
    search_tree_index_path = directory / "search-trees/index.json"
    search_tree_index = json.loads(search_tree_index_path.read_text(encoding="utf-8"))
    for entry in search_tree_index["artifacts"]:
        artifact_path = directory / str(entry["path"])
        fallback_path = directory / str(entry["static_fallback_path"])
        if not artifact_path.is_file():
            raise ReleaseValidationError(
                f"search-tree index references missing artifact: {entry['path']}"
            )
        if not fallback_path.is_file():
            raise ReleaseValidationError(
                "search-tree index references missing static fallback: "
                f"{entry['static_fallback_path']}"
            )
    trace_index = json.loads(trace_index_path.read_text(encoding="utf-8"))
    for entry in trace_index["traces"]:
        trace_path = directory / "traces" / str(entry["path"])
        if not trace_path.is_file():
            raise ReleaseValidationError(f"trace index references missing asset: {entry['path']}")


def publish_release(
    staged_directory: Path,
    data_directory: Path,
    runtime_database: Path,
    version_file: Path,
    site_data_directory: Path,
    readme_path: Path,
    readme_content: str,
    bundle_output_directory: Path,
    *,
    source_commit: str,
    tag: str,
) -> ReleaseBundle:
    verify_release_tree(staged_directory)
    manifests = list(staged_directory.glob("*_manifest.json"))
    manifest = json.loads(manifests[0].read_text(encoding="utf-8"))
    version = str(manifest["version"])
    version_lines = version_file.read_text(encoding="utf-8").splitlines()
    if not version_lines:
        raise ReleaseValidationError("DATASET_VERSION is empty")
    current_version = version_lines[0]
    if version == current_version:
        raise ReleaseValidationError("publish requires a new release version")
    recorded_hash = next(
        (line.removeprefix("sha256=") for line in version_lines if line.startswith("sha256=")), ""
    )
    if sha256_file(runtime_database) != recorded_hash:
        raise ReleaseValidationError("runtime hash does not match DATASET_VERSION")
    runtime_result = verify_database(runtime_database)
    if runtime_result.dataset_version != current_version:
        raise ReleaseValidationError("code version does not match runtime database version")
    expected_database = staged_directory / f"{DATASET_STEM.format(version=version)}.sqlite"
    if not expected_database.exists():
        raise ReleaseValidationError("versioned sqlite filename does not match publish version")
    staged_site_data = staged_directory / str(manifest["artifacts"]["site_data"])
    staged_identity = load_dataset_release_identity(
        staged_directory / str(manifest["artifacts"]["release_identity"])
    )
    if staged_identity.dataset_version != version:
        raise ReleaseValidationError("release identity version does not match manifest")

    from optimization_compass.release_catalog import (
        ReleaseCatalogError,
        catalog_entry_from_bundle,
        load_release_catalog,
        merge_catalog_entry,
        validate_release_catalog,
    )

    catalog_path = data_directory / "releases/catalog.json"
    try:
        load_release_catalog(catalog_path)
    except ReleaseCatalogError as error:
        raise ReleaseValidationError(str(error)) from error

    bundle_output = bundle_output_directory.resolve(strict=False)
    repository_root = ROOT.resolve()
    if bundle_output == repository_root or repository_root in bundle_output.parents:
        raise ReleaseValidationError(
            "complete release bundles must be written outside the repository"
        )

    from optimization_compass.release_bundle import build_release_bundle

    bundle = build_release_bundle(
        staged_directory,
        bundle_output_directory,
        source_commit=source_commit,
        tag=tag,
    )
    try:
        catalog_entry = catalog_entry_from_bundle(
            bundle,
            database_sha256=staged_identity.database_sha256,
        )
    except ReleaseCatalogError as error:
        bundle.path.unlink(missing_ok=True)
        raise ReleaseValidationError(str(error)) from error

    data_directory.parent.mkdir(parents=True, exist_ok=True)
    site_data_directory.parent.mkdir(parents=True, exist_ok=True)
    temporary = Path(
        tempfile.mkdtemp(prefix="optimization-compass-publish-", dir=data_directory.parent)
    )
    data_existed = data_directory.exists()
    runtime_existed = runtime_database.exists()
    version_existed = version_file.exists()
    site_data_existed = site_data_directory.exists()
    readme_existed = readme_path.exists()
    published = False
    try:
        prepared = temporary / "prepared"
        backup = temporary / "backup"
        prepared.mkdir()
        backup.mkdir()
        prepared_data = prepared / "data"
        if data_existed:
            shutil.copytree(data_directory, prepared_data)
            shutil.copytree(data_directory, backup / "data")
        else:
            prepared_data.mkdir()
        if runtime_existed:
            shutil.copy2(runtime_database, backup / "knowledge.sqlite")
        if version_existed:
            shutil.copy2(version_file, backup / "DATASET_VERSION")
        if site_data_existed:
            shutil.copytree(site_data_directory, backup / "site-data")
        if readme_existed:
            shutil.copy2(readme_path, backup / "README.md")
        compact_paths = [
            manifests[0],
            *(
                staged_directory / str(manifest["artifacts"][key])
                for key in COMPACT_RELEASE_ARTIFACT_KEYS
            ),
        ]
        for staged_path in compact_paths:
            shutil.copy2(staged_path, prepared_data / staged_path.name)
        prepared_catalog = prepared_data / "releases/catalog.json"
        try:
            merged_catalog = merge_catalog_entry(
                catalog_path,
                catalog_entry,
                prepared_catalog,
            )
            validate_release_catalog(
                merged_catalog,
                expected_current_identity=staged_identity,
            )
        except ReleaseCatalogError as error:
            raise ReleaseValidationError(str(error)) from error
        prepared_runtime = prepared / "knowledge.sqlite"
        shutil.copy2(expected_database, prepared_runtime)
        prepared_version = prepared / "DATASET_VERSION"
        prepared_version.write_text(
            f"{version}\nsha256={sha256_file(prepared_runtime)}\n", encoding="utf-8"
        )
        prepared_site_data = prepared / "site-data"
        shutil.copytree(staged_site_data, prepared_site_data)
        prepared_readme = prepared / "README.md"
        prepared_readme.write_text(readme_content, encoding="utf-8", newline="\n")

        try:
            displaced_data = temporary / "displaced-data"
            if data_existed:
                _atomic_replace(data_directory, displaced_data)
            _atomic_replace(prepared_data, data_directory)
            _atomic_replace(prepared_runtime, runtime_database)
            _atomic_replace(prepared_version, version_file)
            if site_data_existed:
                _atomic_replace(site_data_directory, temporary / "displaced-site-data")
            _atomic_replace(prepared_site_data, site_data_directory)
            _atomic_replace(prepared_readme, readme_path)
            published = True
        except Exception:
            _restore_publish_targets(
                data_directory=data_directory,
                runtime_database=runtime_database,
                version_file=version_file,
                site_data_directory=site_data_directory,
                readme_path=readme_path,
                backup_directory=backup,
                data_existed=data_existed,
                runtime_existed=runtime_existed,
                version_existed=version_existed,
                site_data_existed=site_data_existed,
                readme_existed=readme_existed,
            )
            raise
    finally:
        shutil.rmtree(temporary, ignore_errors=True)
        if not published and bundle.path.exists():
            bundle.path.unlink()
    return bundle


def _atomic_replace(source: Path, target: Path) -> None:
    source.replace(target)


def _restore_publish_targets(
    *,
    data_directory: Path,
    runtime_database: Path,
    version_file: Path,
    site_data_directory: Path,
    readme_path: Path,
    backup_directory: Path,
    data_existed: bool,
    runtime_existed: bool,
    version_existed: bool,
    site_data_existed: bool,
    readme_existed: bool,
) -> None:
    if data_directory.exists():
        if data_directory.is_dir():
            shutil.rmtree(data_directory)
        else:
            data_directory.unlink()
    if data_existed:
        shutil.copytree(backup_directory / "data", data_directory)
    _restore_file(
        target=runtime_database,
        backup=backup_directory / "knowledge.sqlite",
        existed=runtime_existed,
    )
    _restore_file(
        target=version_file,
        backup=backup_directory / "DATASET_VERSION",
        existed=version_existed,
    )
    if site_data_directory.exists():
        shutil.rmtree(site_data_directory)
    if site_data_existed:
        shutil.copytree(backup_directory / "site-data", site_data_directory)
    _restore_file(
        target=readme_path,
        backup=backup_directory / "README.md",
        existed=readme_existed,
    )


def _restore_file(*, target: Path, backup: Path, existed: bool) -> None:
    if target.exists():
        target.unlink()
    if existed:
        shutil.copy2(backup, target)


def read_snapshot(database_path: Path) -> Snapshot:
    connection = sqlite3.connect(database_path)
    try:
        snapshot: Snapshot = {}
        for table_name in _table_names(connection):
            info = connection.execute(f'PRAGMA table_info("{table_name}")').fetchall()
            columns = tuple(Column(str(row[1]), str(row[2]), int(row[5])) for row in info)
            pk_columns = [
                column.name
                for column in sorted(columns, key=lambda item: item.pk_order)
                if column.pk_order
            ]
            order_columns = pk_columns or [column.name for column in columns]
            order_sql = ", ".join(f'"{column}"' for column in order_columns)
            rows = tuple(
                tuple(row)
                for row in connection.execute(
                    f'SELECT * FROM "{table_name}" ORDER BY {order_sql}'
                ).fetchall()
            )
            snapshot[table_name] = TableSnapshot(columns, rows)
        return snapshot
    finally:
        connection.close()


def _verify_pinned_base(base_database: Path) -> None:
    if not base_database.exists():
        raise ReleaseValidationError(f"base database not found: {base_database}")
    actual_hash = sha256_file(base_database)
    if actual_hash != BASE_DATASET_SHA256:
        raise ReleaseValidationError(f"base database hash mismatch: {actual_hash}")
    connection = sqlite3.connect(base_database)
    try:
        version = _dataset_version(connection)
    finally:
        connection.close()
    if version != BASE_DATASET_VERSION:
        raise ReleaseValidationError(f"base database version mismatch: {version}")


def _apply_atlas_metadata(
    database_path: Path,
    migration_path: Path,
    seed_path: Path,
    *,
    target_version: str,
    release_date: str,
) -> None:
    seed = AtlasMetadataSeed.model_validate_json(seed_path.read_text(encoding="utf-8"))
    predicate_seed = PredicateCatalog.model_validate_json(
        DEFAULT_PREDICATE_SEED.read_text(encoding="utf-8")
    )
    problem_seed = ProblemSuiteSeed.model_validate_json(
        DEFAULT_PROBLEM_SEED.read_text(encoding="utf-8")
    )
    connection = sqlite3.connect(database_path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    try:
        connection.executescript(migration_path.read_text(encoding="utf-8"))
        connection.executescript(DEFAULT_COVERAGE_MIGRATION.read_text(encoding="utf-8"))
        connection.executescript(DEFAULT_PREDICATE_MIGRATION.read_text(encoding="utf-8"))
        connection.executescript(DEFAULT_PROBLEM_MIGRATION.read_text(encoding="utf-8"))
        connection.executescript(DEFAULT_SEMANTIC_VIEW_MIGRATION.read_text(encoding="utf-8"))
        connection.executescript(DEFAULT_VERSIONED_CLAIMS_MIGRATION.read_text(encoding="utf-8"))
        connection.executescript(DEFAULT_FAILURE_MODE_MIGRATION.read_text(encoding="utf-8"))
        connection.executescript(DEFAULT_LEARNING_GRAPH_MIGRATION.read_text(encoding="utf-8"))
        connection.executescript(DEFAULT_TRF_DEFAULTS_MIGRATION.read_text(encoding="utf-8"))
        _insert_problem_seed(connection, problem_seed)
        _insert_seed(connection, seed)
        _insert_predicate_seed(connection, predicate_seed)
        insert_versioned_claims_and_contexts(connection, release_date=release_date)
        insert_structured_failure_modes(connection, release_date=release_date)
        _record_target_release(connection, target_version, release_date)
        connection.execute("DELETE FROM release_checks")
        for check in compute_live_checks(connection, release_date):
            connection.execute(
                """
                INSERT INTO release_checks (
                  check_id, check_name, scope, severity, status, observed_value,
                  expected_condition, details, checked_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    check.check_id,
                    check.check_name,
                    check.scope,
                    check.severity,
                    check.status,
                    check.observed_value,
                    check.expected_condition,
                    check.details,
                    check.checked_at,
                ),
            )
        connection.commit()
        connection.execute("VACUUM")
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    result = verify_database(database_path)
    if not result.ok:
        failed_ids = [item.check_id for item in result.live_failures]
        raise ReleaseValidationError(f"staged database failed live checks: {failed_ids}")


def _record_target_release(
    connection: sqlite3.Connection,
    target_version: str,
    release_date: str,
) -> None:
    if target_version == BASE_DATASET_VERSION:
        return
    connection.execute(
        """
        INSERT INTO version_history (
          version, release_date, status, summary, breaking_changes, source_policy, notes
        ) VALUES (?, ?, 'staged', ?, 'none', ?, ?)
        """,
        (
            target_version,
            release_date,
            "Published canonical Trust Region Reflective defaults and guidance.",
            "official documentation/repositories, original papers, trusted textbooks",
            f"Generated deterministically from the pinned v{BASE_DATASET_VERSION} base.",
        ),
    )
    revision_id = "MR_ATLAS_" + target_version.replace(".", "_")
    connection.execute(
        """
        INSERT INTO model_revisions (
          revision_id, trigger_case_ids, issue_found, schema_change, reason, version, date
        ) VALUES (?, NULL, ?, ?, ?, ?, ?)
        """,
        (
            revision_id,
            "A widely used SciPy default was represented only through broad adjacent methods.",
            (
                "Added a canonical Trust Region Reflective method, primary source, "
                "implementation mapping, and API-default metadata."
            ),
            (
                "Separate library default behavior from method recommendation priority, "
                "and connect the dedicated guide to generated search and retrieval."
            ),
            target_version,
            release_date,
        ),
    )


def _insert_seed(connection: sqlite3.Connection, seed: AtlasMetadataSeed) -> None:
    mappings: tuple[tuple[str, list[Any], dict[str, str]], ...] = (
        (
            "view_presets",
            list(seed.view_presets),
            {
                "relation_types": "relation_types_json",
                "filter_policy": "filter_policy_json",
                "focus_fallback_entity_types": "focus_fallback_entity_types_json",
                "source_ids": "source_ids_json",
            },
        ),
        (
            "method_visualization_profiles",
            list(seed.method_visualization_profiles),
            {
                "state_fields": "state_fields_json",
                "event_types": "event_types_json",
                "source_ids": "source_ids_json",
            },
        ),
        (
            "demo_scenarios",
            list(seed.demo_scenarios),
            {
                "initial_point": "initial_point_json",
                "parameters": "parameters_json",
                "stopping": "stopping_json",
                "source_ids": "source_ids_json",
            },
        ),
        (
            "comparison_sets",
            list(seed.comparison_sets),
            {
                "initial_point": "initial_point_json",
                "stopping": "stopping_json",
                "source_ids": "source_ids_json",
            },
        ),
        (
            "comparison_set_members",
            list(seed.comparison_set_members),
            {"parameters": "parameters_json"},
        ),
        ("learning_edges", list(seed.learning_edges), {"source_ids": "source_ids_json"}),
        (
            "terminology_aliases",
            list(seed.terminology_aliases),
            {
                "abbreviations": "abbreviations_json",
                "synonyms": "synonyms_json",
                "domain_terms": "domain_terms_json",
                "misspellings": "misspellings_json",
                "deprecated_terms": "deprecated_terms_json",
                "source_ids": "source_ids_json",
            },
        ),
        (
            "learning_slice_priorities",
            list(seed.learning_slice_priorities),
            {"source_ids": "source_ids_json"},
        ),
        (
            "learning_coverage_expectations",
            list(seed.learning_coverage_expectations),
            {"source_ids": "source_ids_json"},
        ),
    )
    for table_name, models, json_columns in mappings:
        for model in models:
            values = model.model_dump()
            for source_name, target_name in json_columns.items():
                values[target_name] = _canonical_json(values.pop(source_name))
            columns = list(values)
            placeholders = ", ".join("?" for _ in columns)
            column_sql = ", ".join(f'"{column}"' for column in columns)
            connection.execute(
                f'INSERT INTO "{table_name}" ({column_sql}) VALUES ({placeholders})',
                [values[column] for column in columns],
            )


def _insert_problem_seed(connection: sqlite3.Connection, seed: ProblemSuiteSeed) -> None:
    for definition in seed.definitions:
        values = definition.model_dump(mode="json")
        related_problem_ids = values.pop("related_problem_ids")
        feature_ids = values.pop("feature_ids")
        values["available_oracles_json"] = _canonical_json(values.pop("available_oracles"))
        values["dimensionality_policy_json"] = _canonical_json(values.pop("dimensionality_policy"))
        values["related_problem_ids_json"] = _canonical_json(related_problem_ids)
        values["feature_ids_json"] = _canonical_json(feature_ids)
        values["source_ids_json"] = _canonical_json(values.pop("source_ids"))
        _insert_mapping(connection, "problem_definitions", values)
        for problem_id in related_problem_ids:
            _insert_mapping(
                connection,
                "problem_definition_archetypes",
                {
                    "problem_definition_id": definition.problem_definition_id,
                    "problem_id": problem_id,
                },
            )
        for feature_id in feature_ids:
            _insert_mapping(
                connection,
                "problem_definition_features",
                {
                    "problem_definition_id": definition.problem_definition_id,
                    "feature_id": feature_id,
                },
            )

    for instance in seed.instances:
        values = instance.model_dump(mode="json")
        for source_name in (
            "parameters",
            "bounds",
            "constraints",
            "initialization_candidates",
            "display",
            "intended_phenomena",
        ):
            values[f"{source_name}_json"] = _canonical_json(values.pop(source_name))
        reference = values.pop("known_reference")
        values["known_reference_json"] = (
            _canonical_json(reference) if reference is not None else None
        )
        values["source_ids_json"] = _canonical_json(values.pop("source_ids"))
        _insert_mapping(connection, "problem_instances", values)


def _insert_predicate_seed(connection: sqlite3.Connection, seed: PredicateCatalog) -> None:
    for predicate in seed.predicates:
        values = predicate.model_dump(mode="json")
        values["value_json"] = _canonical_json(values.pop("value"))
        values["source_ids_json"] = _canonical_json(values.pop("source_ids"))
        _insert_mapping(connection, "atomic_predicates", values)

    for policy in seed.policies:
        values = policy.model_dump(mode="json")
        expression = values.pop("expression")
        values["expression_json"] = _canonical_json(expression) if expression is not None else None
        values["source_ids_json"] = _canonical_json(values.pop("source_ids"))
        _insert_mapping(connection, "predicate_policies", values)

    for coverage in seed.coverage:
        values = coverage.model_dump(mode="json")
        values["source_ids_json"] = _canonical_json(values.pop("source_ids"))
        _insert_mapping(connection, "predicate_coverage", values)

    for retirement in seed.rule_target_retirements:
        values = retirement.model_dump(mode="json")
        values["source_ids_json"] = _canonical_json(values.pop("source_ids"))
        _insert_mapping(connection, "decision_rule_target_retirements", values)


def _insert_mapping(
    connection: sqlite3.Connection, table_name: str, values: dict[str, Any]
) -> None:
    columns = list(values)
    placeholders = ", ".join("?" for _ in columns)
    column_sql = ", ".join(f'"{column}"' for column in columns)
    connection.execute(
        f'INSERT INTO "{table_name}" ({column_sql}) VALUES ({placeholders})',
        [values[column] for column in columns],
    )


def _write_ddl(database_path: Path, destination: Path) -> None:
    connection = sqlite3.connect(database_path)
    try:
        rows = connection.execute(
            """
            SELECT type, name, sql
            FROM sqlite_master
            WHERE sql IS NOT NULL AND name NOT LIKE 'sqlite_%'
            ORDER BY CASE type WHEN 'table' THEN 0 WHEN 'index' THEN 1 ELSE 2 END, name
            """
        ).fetchall()
    finally:
        connection.close()
    statements = ["PRAGMA foreign_keys = ON;"]
    statements.extend(f"{str(row[2]).rstrip(';')};" for row in rows)
    destination.write_text("\n\n".join(statements) + "\n", encoding="utf-8")


def _write_json(
    snapshot: Snapshot,
    destination: Path,
    *,
    version: str,
    release_date: str,
) -> None:
    payload = {
        "version": version,
        "release_date": release_date,
        "schemas": _schema_payload(snapshot),
        "tables": {
            name: [_row_dict(table, row) for row in table.rows] for name, table in snapshot.items()
        },
    }
    destination.write_text(_canonical_json(payload, pretty=True), encoding="utf-8")


def _write_jsonl(
    snapshot: Snapshot,
    destination: Path,
    *,
    version: str,
    release_date: str,
) -> None:
    lines = [
        _canonical_json(
            {
                "type": "release",
                "version": version,
                "release_date": release_date,
                "schemas": _schema_payload(snapshot),
            }
        )
    ]
    for name, table in snapshot.items():
        lines.extend(
            _canonical_json({"table": name, "row": _row_dict(table, row)}) for row in table.rows
        )
    destination.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_csv_directory(snapshot: Snapshot, destination: Path) -> None:
    destination.mkdir()
    for name, table in snapshot.items():
        with (destination / f"{name}.csv").open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle, lineterminator="\n")
            writer.writerow([column.name for column in table.columns])
            writer.writerows([_encode_cell(value) for value in row] for row in table.rows)


def _write_csv_zip(csv_directory: Path, destination: Path, *, release_date: str) -> None:
    zip_time = _zip_time(release_date)
    with zipfile.ZipFile(
        destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as archive:
        for path in sorted(csv_directory.glob("*.csv")):
            info = zipfile.ZipInfo(path.name, date_time=zip_time)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o100644 << 16
            archive.writestr(
                info, path.read_bytes(), compress_type=zipfile.ZIP_DEFLATED, compresslevel=9
            )
        license_directory = csv_directory.parent / "licenses"
        for name in DATA_ZIP_LICENSES:
            path = license_directory / name
            info = zipfile.ZipInfo(f"LICENSES/{name}", date_time=zip_time)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o100644 << 16
            archive.writestr(
                info, path.read_bytes(), compress_type=zipfile.ZIP_DEFLATED, compresslevel=9
            )


def _write_xlsx(snapshot: Snapshot, destination: Path, *, release_date: str) -> None:
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is None:
        raise ReleaseValidationError("new workbook did not create an active sheet")
    workbook.remove(active_sheet)
    fixed = _release_datetime(release_date)
    workbook.properties.created = fixed
    workbook.properties.modified = fixed
    for name, table in snapshot.items():
        worksheet = workbook.create_sheet(name)
        worksheet.append([column.name for column in table.columns])
        for row in table.rows:
            worksheet.append([_encode_cell(value) for value in row])
        worksheet.freeze_panes = "A2"
    workbook.save(destination)
    _normalize_zip(destination, release_date=release_date)


def _write_report(
    snapshot: Snapshot,
    destination: Path,
    *,
    version: str,
    release_date: str,
) -> None:
    lines = [
        "# Optimization Method Selection Database staged report",
        "",
        f"- Version: `{version}`",
        f"- Release date: `{release_date}`",
        f"- Tables: `{len(snapshot)}`",
        f"- Rows: `{sum(len(table.rows) for table in snapshot.values())}`",
        "- Data license: `CC-BY-4.0` (`licenses/DATA_LICENSE.txt`)",
        "",
        "| Table | Rows |",
        "|---|---:|",
    ]
    lines.extend(f"| `{name}` | {len(table.rows)} |" for name, table in snapshot.items())
    free_text_conditions = _free_text_only_method_conditions(snapshot)
    affected_methods = sorted({method_id for method_id, _ in free_text_conditions})
    lines.extend(
        [
            "",
            "## Free-text-only method conditions",
            "",
            (
                f"Detected `{len(free_text_conditions)}` populated condition fields across "
                f"`{len(affected_methods)}` methods without complete atomic-predicate coverage."
            ),
            "",
            "| Method | Unmigrated condition fields |",
            "|---|---|",
        ]
    )
    fields_by_method: dict[str, list[str]] = {}
    for method_id, field_name in free_text_conditions:
        fields_by_method.setdefault(method_id, []).append(field_name)
    lines.extend(
        f"| `{method_id}` | {', '.join(f'`{field}`' for field in fields_by_method[method_id])} |"
        for method_id in affected_methods
    )
    destination.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _free_text_only_method_conditions(snapshot: Snapshot) -> list[tuple[str, str]]:
    methods = snapshot.get("methods")
    coverage = snapshot.get("predicate_coverage")
    if methods is None or coverage is None:
        return []
    complete_methods: set[str] = set()
    coverage_columns = {column.name: index for index, column in enumerate(coverage.columns)}
    for row in coverage.rows:
        if (
            row[coverage_columns["subject_type"]] == "method"
            and row[coverage_columns["status"]] == "complete"
        ):
            complete_methods.add(str(row[coverage_columns["subject_id"]]))

    method_columns = {column.name: index for index, column in enumerate(methods.columns)}
    condition_fields = (
        "required_assumptions",
        "avoid_conditions",
        "first_choice_conditions",
        "second_choice_conditions",
    )
    result: list[tuple[str, str]] = []
    for row in methods.rows:
        method_id = str(row[method_columns["method_id"]])
        if method_id in complete_methods:
            continue
        for field_name in condition_fields:
            value = row[method_columns[field_name]]
            if value is not None and str(value).strip():
                result.append((method_id, field_name))
    return result


def _manifest_payload(
    directory: Path,
    stem: str,
    database_path: Path,
    snapshot: Snapshot,
    *,
    version: str,
    release_date: str,
    include_manifest: bool,
) -> dict[str, Any]:
    manifest_name = f"{stem}_manifest.json"
    excluded = set() if include_manifest else {manifest_name}
    return {
        "schema_version": 2,
        "version": version,
        "release_date": release_date,
        "base_sha256": BASE_DATASET_SHA256,
        "database_sha256": sha256_file(database_path),
        "artifacts": {
            "database": f"{stem}.sqlite",
            "ddl": f"{stem}_schema.sql",
            "json": f"{stem}.json",
            "jsonl": f"{stem}.jsonl",
            "csv_directory": f"{stem}_csv",
            "csv_zip": f"{stem}_csv.zip",
            "xlsx": f"{stem}.xlsx",
            "report": f"{stem}_report.md",
            "release_identity": f"{stem}_release.json",
            "site_data": f"{stem}_site-data",
        },
        "files": _artifact_hashes(directory, exclude=excluded),
        "table_counts": {name: len(table.rows) for name, table in snapshot.items()},
        "licenses": RELEASE_LICENSE_MANIFEST,
        "validation": {
            "formats": ["csv_directory", "csv_zip", "ddl", "json", "jsonl", "sqlite", "xlsx"],
            "live_release_checks": "pass",
            "round_trip": "pass",
        },
    }


def _artifact_hashes(directory: Path, *, exclude: set[str]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for path in sorted(item for item in directory.rglob("*") if item.is_file()):
        relative = path.relative_to(directory).as_posix()
        if relative in exclude:
            continue
        result[relative] = {"bytes": path.stat().st_size, "sha256": sha256_file(path)}
    return result


def _read_json(path: Path) -> Snapshot:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return _snapshot_from_row_dicts(payload["tables"], payload["schemas"])


def _read_jsonl(path: Path) -> Snapshot:
    grouped: dict[str, list[dict[str, Any]]] = {}
    schemas: dict[str, list[dict[str, Any]]] | None = None
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines()):
        payload = json.loads(line)
        if line_number == 0:
            schemas = payload["schemas"]
            continue
        grouped.setdefault(str(payload["table"]), []).append(payload["row"])
    if schemas is None:
        raise ReleaseValidationError("jsonl release header is missing schemas")
    return _snapshot_from_row_dicts(grouped, schemas)


def _read_csv_directory(directory: Path, reference: Snapshot) -> Snapshot:
    result: Snapshot = {}
    for name, table in reference.items():
        with (directory / f"{name}.csv").open(encoding="utf-8", newline="") as handle:
            rows = list(csv.reader(handle))
        _assert_header(name, rows[0], table)
        result[name] = TableSnapshot(
            table.columns,
            tuple(tuple(_decode_cell(value) for value in row) for row in rows[1:]),
        )
    return result


def _read_csv_zip(path: Path, reference: Snapshot) -> Snapshot:
    result: Snapshot = {}
    with zipfile.ZipFile(path) as archive:
        expected_entries = [f"{name}.csv" for name in reference] + [
            f"LICENSES/{name}" for name in DATA_ZIP_LICENSES
        ]
        if archive.namelist() != expected_entries:
            raise ReleaseValidationError("csv zip entry order/content differs from table order")
        for name, table in reference.items():
            text = archive.read(f"{name}.csv").decode("utf-8")
            rows = list(csv.reader(io.StringIO(text)))
            _assert_header(name, rows[0], table)
            result[name] = TableSnapshot(
                table.columns,
                tuple(tuple(_decode_cell(value) for value in row) for row in rows[1:]),
            )
        license_directory = path.parent / "licenses"
        for name in DATA_ZIP_LICENSES:
            if archive.read(f"LICENSES/{name}") != (license_directory / name).read_bytes():
                raise ReleaseValidationError(f"csv zip license differs from release notice: {name}")
    return result


def _read_xlsx(path: Path, reference: Snapshot) -> Snapshot:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != list(reference):
            raise ReleaseValidationError("xlsx sheet order/content differs from table order")
        result: Snapshot = {}
        for name, table in reference.items():
            worksheet = workbook[name]
            rows = list(worksheet.iter_rows(values_only=True))
            header = [str(value) for value in rows[0]]
            _assert_header(name, header, table)
            result[name] = TableSnapshot(
                table.columns,
                tuple(tuple(_decode_cell(str(value)) for value in row) for row in rows[1:]),
            )
        return result
    finally:
        workbook.close()


def _verify_ddl(path: Path, reference: Snapshot) -> None:
    connection = sqlite3.connect(":memory:")
    try:
        connection.executescript(path.read_text(encoding="utf-8"))
        observed = {
            name: tuple(
                Column(str(row[1]), str(row[2]), int(row[5]))
                for row in connection.execute(f'PRAGMA table_info("{name}")').fetchall()
            )
            for name in _table_names(connection)
        }
    finally:
        connection.close()
    expected = {name: table.columns for name, table in reference.items()}
    if observed != expected:
        raise ReleaseValidationError("ddl schema differs from sqlite")


def _snapshot_from_row_dicts(
    tables: dict[str, list[dict[str, Any]]],
    schemas: dict[str, list[dict[str, Any]]],
) -> Snapshot:
    result: Snapshot = {}
    for name, rows in tables.items():
        columns = tuple(
            Column(str(column["name"]), str(column["declared_type"]), int(column["pk_order"]))
            for column in schemas[name]
        )
        result[name] = TableSnapshot(
            columns, tuple(tuple(row[column.name] for column in columns) for row in rows)
        )
    return result


def _schema_payload(snapshot: Snapshot) -> dict[str, list[dict[str, Any]]]:
    return {
        name: [
            {
                "name": column.name,
                "declared_type": column.declared_type,
                "pk_order": column.pk_order,
            }
            for column in table.columns
        ]
        for name, table in snapshot.items()
    }


def _row_dict(table: TableSnapshot, row: tuple[Any, ...]) -> dict[str, Any]:
    return {column.name: value for column, value in zip(table.columns, row, strict=True)}


def _encode_cell(value: Any) -> str:
    if value is None:
        return "\\N"
    if isinstance(value, bytes):
        return "\\B" + value.hex()
    if isinstance(value, bool):
        return "\\I" + str(int(value))
    if isinstance(value, int):
        return "\\I" + str(value)
    if isinstance(value, float):
        return "\\F" + value.hex()
    return "\\S" + str(value)


def _decode_cell(value: str) -> Any:
    prefix, payload = value[:2], value[2:]
    if value == "\\N":
        return None
    if prefix == "\\B":
        return bytes.fromhex(payload)
    if prefix == "\\I":
        return int(payload)
    if prefix == "\\F":
        return float.fromhex(payload)
    if prefix == "\\S":
        return payload
    raise ReleaseValidationError(f"invalid typed cell: {value[:20]}")


def _release_datetime(release_date: str) -> datetime:
    return datetime.strptime(release_date, "%Y-%m-%d").replace(tzinfo=UTC)


def _zip_time(release_date: str) -> tuple[int, int, int, int, int, int]:
    value = _release_datetime(release_date)
    return (value.year, value.month, value.day, 0, 0, 0)


def _normalize_zip(path: Path, *, release_date: str) -> None:
    zip_time = _zip_time(release_date)
    modified = f"{release_date}T00:00:00Z".encode()
    with zipfile.ZipFile(path, "r") as source:
        entries = [
            (name, source.read(name), source.getinfo(name).compress_type)
            for name in sorted(source.namelist())
        ]
    temporary = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(temporary, "w") as destination:
        for name, content, compression in entries:
            if name == "docProps/core.xml":
                content = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>" + modified + rb"\g<2>",
                    content,
                )
            info = zipfile.ZipInfo(name, date_time=zip_time)
            info.compress_type = compression
            info.external_attr = 0o100644 << 16
            destination.writestr(info, content, compress_type=compression, compresslevel=9)
    temporary.replace(path)


def _assert_header(name: str, header: list[str], table: TableSnapshot) -> None:
    if header != [column.name for column in table.columns]:
        raise ReleaseValidationError(f"{name} columns differ from sqlite")


def _canonical_json(value: Any, *, pretty: bool = False) -> str:
    if pretty:
        return json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)


def _table_names(connection: sqlite3.Connection) -> list[str]:
    return [
        str(row[0])
        for row in connection.execute(
            "SELECT name FROM sqlite_master "
            "WHERE type = 'table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
        ).fetchall()
    ]


def _dataset_version(connection: sqlite3.Connection) -> str:
    row = connection.execute(
        "SELECT version FROM version_history ORDER BY release_date DESC, version DESC LIMIT 1"
    ).fetchone()
    return str(row[0]) if row else "unknown"


def _release_date(connection: sqlite3.Connection) -> str:
    row = connection.execute(
        "SELECT release_date FROM version_history "
        "WHERE release_date IS NOT NULL "
        "ORDER BY release_date DESC, version DESC LIMIT 1"
    ).fetchone()
    if row is None:
        raise ReleaseValidationError("version_history has no release date")
    return str(row[0])


def _stored_checks(connection: sqlite3.Connection) -> list[LiveCheck]:
    if "release_checks" not in _table_names(connection):
        return []
    return [
        LiveCheck(
            check_id=str(row["check_id"]),
            check_name=str(row["check_name"]),
            scope=str(row["scope"]),
            severity=str(row["severity"]),
            status=row["status"],
            observed_value=str(row["observed_value"] or ""),
            expected_condition=str(row["expected_condition"]),
            details=str(row["details"] or ""),
            checked_at=str(row["checked_at"]),
        )
        for row in connection.execute("SELECT * FROM release_checks ORDER BY check_id").fetchall()
    ]


def _check(
    check_id: str,
    name: str,
    scope: str,
    severity: str,
    passed: bool,
    observed: str,
    expected: str,
    details: str,
    checked_at: str,
    *,
    warn: bool = False,
) -> LiveCheck:
    status: Literal["pass", "warn", "fail", "not_run"] = "pass" if passed else "fail"
    if passed and warn:
        status = "warn"
    return LiveCheck(
        check_id, name, scope, severity, status, observed, expected, details, checked_at
    )


def _duplicate_primary_keys(connection: sqlite3.Connection, tables: list[str]) -> list[str]:
    issues: list[str] = []
    for table in tables:
        pk = [
            str(row[1])
            for row in sorted(
                (row for row in connection.execute(f'PRAGMA table_info("{table}")') if row[5]),
                key=lambda row: row[5],
            )
        ]
        if not pk:
            continue
        columns = ", ".join(f'"{name}"' for name in pk)
        row = connection.execute(
            f'SELECT 1 FROM "{table}" GROUP BY {columns} HAVING COUNT(*) > 1 LIMIT 1'
        ).fetchone()
        if row:
            issues.append(table)
    return issues


def _missing_source_references(connection: sqlite3.Connection, tables: list[str]) -> list[str]:
    if "sources" not in tables:
        return ["sources table"]
    known = {str(row[0]) for row in connection.execute("SELECT source_id FROM sources")}
    issues: set[str] = set()
    for table in tables:
        columns = [str(row[1]) for row in connection.execute(f'PRAGMA table_info("{table}")')]
        for column in columns:
            if table == "learning_edges" and column == "source_id":
                continue
            if not (
                column == "source_id"
                or column.endswith("source_ids")
                or column.endswith("source_ids_json")
            ):
                continue
            for (raw,) in connection.execute(
                f'SELECT "{column}" FROM "{table}" WHERE "{column}" IS NOT NULL'
            ):
                for source_id in _parse_ids(raw):
                    if source_id not in known:
                        issues.add(f"{table}.{column}:{source_id}")
    return sorted(issues)


def _parse_ids(raw: Any) -> list[str]:
    text = str(raw).strip()
    if not text:
        return []
    if text.startswith("["):
        value = json.loads(text)
        return [str(item) for item in value]
    return [item.strip() for item in text.split(";") if item.strip()]


def _unresolved_evidence_targets(connection: sqlite3.Connection, tables: list[str]) -> list[str]:
    if "evidence_links" not in tables:
        return []
    issues: list[str] = []
    table_set = set(tables)
    for link_id, target_table, target_id in connection.execute(
        "SELECT evidence_link_id, target_table, target_id FROM evidence_links"
    ):
        if target_table not in table_set:
            issues.append(str(link_id))
            continue
        pk = [row for row in connection.execute(f'PRAGMA table_info("{target_table}")') if row[5]]
        if not pk:
            issues.append(str(link_id))
            continue
        first_pk = str(sorted(pk, key=lambda row: row[5])[0][1])
        found = connection.execute(
            f'SELECT 1 FROM "{target_table}" WHERE "{first_pk}" = ? LIMIT 1', (target_id,)
        ).fetchone()
        if found is None:
            issues.append(str(link_id))
    return issues


def _unresolved_decision_targets(connection: sqlite3.Connection) -> list[str]:
    mapping = {
        "alternative": ("alternative_solution_checks", "alternative_id"),
        "feature": ("problem_features", "feature_id"),
        "method": ("methods", "method_id"),
        "problem": ("problem_archetypes", "problem_id"),
    }
    issues: list[str] = []
    for rule_id, target_type, raw_ids in connection.execute(
        "SELECT rule_id, action_target_type, action_target_ids FROM decision_rules"
    ):
        if target_type == "none":
            if _parse_ids(raw_ids):
                issues.append(str(rule_id))
            continue
        target = mapping.get(str(target_type))
        if target is None:
            issues.append(str(rule_id))
            continue
        table, column = target
        for target_id in _parse_ids(raw_ids):
            if (
                connection.execute(
                    f'SELECT 1 FROM "{table}" WHERE "{column}" = ?', (target_id,)
                ).fetchone()
                is None
            ):
                issues.append(f"{rule_id}:{target_id}")
    return issues


def _blank_foreign_keys(connection: sqlite3.Connection, tables: list[str]) -> list[str]:
    issues: list[str] = []
    for table in tables:
        foreign_columns = {
            str(row[3]) for row in connection.execute(f'PRAGMA foreign_key_list("{table}")')
        }
        for column in foreign_columns:
            count = int(
                connection.execute(
                    f'SELECT COUNT(*) FROM "{table}" '
                    f"WHERE typeof(\"{column}\") = 'text' AND trim(\"{column}\") = ''"
                ).fetchone()[0]
            )
            if count:
                issues.append(f"{table}.{column}:{count}")
    return issues


def _objective_form_issues(connection: sqlite3.Connection) -> list[str]:
    issues: list[str] = []
    has_value = connection.execute(
        "SELECT 1 FROM feature_values WHERE feature_value_id = 'FV0093'"
    ).fetchone()
    if has_value is None:
        issues.append("FV0093 missing")
    count = int(
        connection.execute(
            "SELECT COUNT(*) FROM case_feature_map "
            "WHERE lower(coalesce(value_text, '')) = 'discrete'"
        ).fetchone()[0]
    )
    if count:
        issues.append(f"free-text discrete:{count}")
    return issues


def _implementation_release_coverage(connection: sqlite3.Connection) -> tuple[int, int]:
    row = connection.execute(
        "SELECT SUM(CASE WHEN last_release IS NOT NULL "
        "AND trim(last_release) <> '' AND lower(last_release) <> 'unknown' "
        "THEN 1 ELSE 0 END), COUNT(*) FROM implementations"
    ).fetchone()
    return int(row[0] or 0), int(row[1])


def _license_issues(connection: sqlite3.Connection) -> list[str]:
    expectations = {"I_MANOPT_MATLAB": "GPL-3.0-or-later", "I_NOMAD": "LGPL-3.0-or-later"}
    issues: list[str] = []
    for implementation_id, expected in expectations.items():
        row = connection.execute(
            "SELECT license FROM implementations WHERE implementation_id = ?", (implementation_id,)
        ).fetchone()
        if row is None or row[0] != expected:
            issues.append(implementation_id)
    return issues


def _maintenance_issues(connection: sqlite3.Connection) -> list[str]:
    row = connection.execute(
        "SELECT maintenance_status, last_release FROM implementations "
        "WHERE implementation_id = 'I_JAXOPT'"
    ).fetchone()
    return [] if row is not None and tuple(row) == ("legacy", "0.8.5") else ["I_JAXOPT"]


def _json_array_has_duplicates(raw: str) -> bool:
    try:
        values = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return True
    return not isinstance(values, list) or not values or len(values) != len(set(values))


def _valid_json_container(
    raw: Any, expected_type: type[object], *, non_empty: bool = False
) -> bool:
    try:
        value = json.loads(str(raw))
    except (json.JSONDecodeError, TypeError):
        return False
    return isinstance(value, expected_type) and (not non_empty or bool(value))


def _missing_table_issues(tables: list[str], required: set[str]) -> list[str]:
    return [f"missing:{table}" for table in sorted(required - set(tables))]


def _view_preset_issues(connection: sqlite3.Connection) -> list[str]:
    issues: list[str] = []
    root_targets = {
        "problem": ("problem_archetypes", "problem_id"),
        "method": ("methods", "method_id"),
        "view_preset": ("view_presets", "preset_id"),
    }
    columns = {str(row[1]) for row in connection.execute('PRAGMA table_info("view_presets")')}
    if "view_id" not in columns:
        for preset_id, status, root_type, root_id, relations in connection.execute(
            "SELECT preset_id, root_support_status, root_entity_type, root_entity_id, "
            "relation_types_json FROM view_presets"
        ):
            if (status == "supported") != (root_type is not None and root_id is not None):
                issues.append(str(preset_id))
            if status == "supported":
                resolver = root_targets.get(str(root_type))
                if resolver is None:
                    issues.append(f"{preset_id}:invalid-root-type")
                else:
                    table, column = resolver
                    if (
                        connection.execute(
                            f'SELECT 1 FROM "{table}" WHERE "{column}" = ?', (root_id,)
                        ).fetchone()
                        is None
                    ):
                        issues.append(f"{preset_id}:unresolved-root")
            if _json_array_has_duplicates(str(relations)):
                issues.append(f"{preset_id}:duplicate-relations")
        return issues
    view_ids: set[str] = set()
    for (
        preset_id,
        view_id,
        status,
        root_type,
        root_id,
        relations,
        filters,
        fallback_types,
    ) in connection.execute(
        "SELECT preset_id, view_id, root_support_status, root_entity_type, root_entity_id, "
        "relation_types_json, filter_policy_json, focus_fallback_entity_types_json "
        "FROM view_presets"
    ):
        if str(view_id) in view_ids:
            issues.append(f"{preset_id}:duplicate-view-id")
        view_ids.add(str(view_id))
        if (status == "supported") != (root_type is not None and root_id is not None):
            issues.append(str(preset_id))
        if status == "supported":
            resolver = root_targets.get(str(root_type))
            if resolver is None:
                issues.append(f"{preset_id}:invalid-root-type")
                continue
            table, column = resolver
            if (
                connection.execute(
                    f'SELECT 1 FROM "{table}" WHERE "{column}" = ?', (root_id,)
                ).fetchone()
                is None
            ):
                issues.append(f"{preset_id}:unresolved-root")
        if _json_array_has_duplicates(str(relations)):
            issues.append(f"{preset_id}:duplicate-relations")
        if not _valid_json_container(filters, dict, non_empty=True):
            issues.append(f"{preset_id}:invalid-filter-policy")
        else:
            filter_policy = json.loads(str(filters))
            groups = filter_policy.get("groups")
            if filter_policy.get("mode") != "authored_groups" or not isinstance(groups, list):
                issues.append(f"{preset_id}:invalid-filter-policy")
            else:
                seen_groups: set[str] = set()
                seen_selectors: set[tuple[str, str]] = set()
                selector_tables = {
                    "question_ids": ("decision_questions", "question_id"),
                    "feature_ids": ("problem_features", "feature_id"),
                    "method_ids": ("methods", "method_id"),
                    "alternative_ids": (
                        "alternative_solution_checks",
                        "alternative_id",
                    ),
                }
                for group in groups:
                    if not isinstance(group, dict):
                        issues.append(f"{preset_id}:invalid-filter-group")
                        continue
                    group_id = str(group.get("group_id") or "")
                    if (
                        not group_id.strip()
                        or group_id in seen_groups
                        or not str(group.get("label_ja") or "").strip()
                        or not str(group.get("label_en") or "").strip()
                    ):
                        issues.append(f"{preset_id}:invalid-filter-group")
                    seen_groups.add(group_id)
                    selector_count = 0
                    for key, (table, column) in selector_tables.items():
                        values = group.get(key, [])
                        if not isinstance(values, list) or len(values) != len(set(values)):
                            issues.append(f"{preset_id}:invalid-filter-selectors")
                            continue
                        selector_count += len(values)
                        for raw_id in values:
                            selector = (key, str(raw_id))
                            if selector in seen_selectors:
                                issues.append(f"{preset_id}:duplicate-filter-selector")
                            seen_selectors.add(selector)
                            if (
                                connection.execute(
                                    f'SELECT 1 FROM "{table}" WHERE "{column}" = ?',
                                    (raw_id,),
                                ).fetchone()
                                is None
                            ):
                                issues.append(f"{preset_id}:unresolved-filter-selector")
                    if selector_count == 0:
                        issues.append(f"{preset_id}:empty-filter-group")
        if not _valid_json_container(fallback_types, list, non_empty=True):
            issues.append(f"{preset_id}:invalid-focus-fallback")
    return issues


def _profile_objective_issues(connection: sqlite3.Connection) -> list[str]:
    issues: list[str] = []
    for row in connection.execute(
        """
        SELECT profile.profile_id, profile.support_status, profile.min_dimension,
               profile.max_dimension, profile.generator_id,
               profile.implementation_status, profile.implementation_id,
               profile.state_fields_json, profile.event_types_json,
               method.method_id, implementation.implementation_id
        FROM method_visualization_profiles AS profile
        LEFT JOIN methods AS method USING (method_id)
        LEFT JOIN implementations AS implementation
          ON profile.implementation_id = implementation.implementation_id
        """
    ):
        profile_id = str(row[0])
        if row[9] is None or row[1] not in {
            "supported",
            "unsupported",
            "unknown",
            "not_applicable",
        }:
            issues.append(f"{profile_id}:method-or-support")
        if int(row[2]) < 1 or int(row[3]) < int(row[2]) or not str(row[4]).strip():
            issues.append(f"{profile_id}:dimension-or-generator")
        if (row[5] == "supported") != (row[6] is not None and row[10] is not None):
            issues.append(f"{profile_id}:implementation")
        if _json_array_has_duplicates(str(row[7])) or _json_array_has_duplicates(str(row[8])):
            issues.append(f"{profile_id}:state-or-event-json")
    for row in connection.execute(
        """
        SELECT problem_definition_id, objective_direction, available_oracles_json,
               dimensionality_policy_json, related_problem_ids_json, feature_ids_json,
               source_ids_json
        FROM problem_definitions
        """
    ):
        valid = (
            row[1] in {"minimize", "maximize", "multiobjective"}
            and _valid_json_container(row[2], list, non_empty=True)
            and _valid_json_container(row[3], dict, non_empty=True)
            and _valid_json_container(row[4], list, non_empty=True)
            and _valid_json_container(row[5], list, non_empty=True)
            and _valid_json_container(row[6], list, non_empty=True)
        )
        if not valid:
            issues.append(f"{row[0]}:definition-metadata")
    for row in connection.execute(
        """
        SELECT problem_instance_id, registry_key, dimension, parameters_json,
               bounds_json, constraints_json, initialization_candidates_json,
               seed_status, seed_value, known_reference_status, known_reference_json,
               display_json, intended_phenomena_json, source_ids_json
        FROM problem_instances
        """
    ):
        reference_required = row[9] in {"known_exact", "known_reference", "best_known"}
        valid = (
            bool(str(row[1]).strip())
            and int(row[2]) >= 1
            and _valid_json_container(row[3], dict, non_empty=True)
            and _valid_json_container(row[4], dict, non_empty=True)
            and _valid_json_container(row[5], list)
            and _valid_json_container(row[6], list, non_empty=True)
            and row[7] in {"fixed", "not_applicable", "unknown"}
            and ((row[7] == "fixed") == (row[8] is not None))
            and row[9]
            in {"known_exact", "known_reference", "best_known", "unknown", "not_meaningful"}
            and (reference_required == (row[10] is not None))
            and _valid_json_container(row[11], dict, non_empty=True)
            and _valid_json_container(row[12], list, non_empty=True)
            and _valid_json_container(row[13], list, non_empty=True)
        )
        if not valid:
            issues.append(f"{row[0]}:instance-metadata")
    return issues


def _scenario_issues(connection: sqlite3.Connection) -> list[str]:
    issues: list[str] = []
    for row in connection.execute(
        """
        SELECT scenario.scenario_id, scenario.seed_status, scenario.seed_value,
               scenario.budget, scenario.initial_point_json, scenario.parameters_json,
               scenario.stopping_json, profile.profile_id, instance.problem_instance_id
        FROM demo_scenarios AS scenario
        LEFT JOIN method_visualization_profiles AS profile
          ON scenario.method_id = profile.method_id AND scenario.profile_id = profile.profile_id
        LEFT JOIN problem_instances AS instance USING (problem_instance_id)
        """
    ):
        valid = (
            row[1] in {"fixed", "not_applicable", "unknown"}
            and ((row[1] == "fixed") == (row[2] is not None))
            and int(row[3]) > 0
            and _valid_json_container(row[4], list, non_empty=True)
            and _valid_json_container(row[5], dict)
            and _valid_json_container(row[6], dict)
            and row[7] is not None
            and row[8] is not None
        )
        if not valid:
            issues.append(str(row[0]))
    return issues


def _comparison_issues(connection: sqlite3.Connection) -> list[str]:
    issues: list[str] = []
    for row in connection.execute(
        """
        SELECT comparison.comparison_set_id, comparison.synchronization,
               comparison.fairness_note, comparison.seed_status, comparison.seed_value,
               comparison.budget, comparison.initial_point_json, comparison.stopping_json,
               instance.problem_instance_id
        FROM comparison_sets AS comparison
        LEFT JOIN problem_instances AS instance USING (problem_instance_id)
        """
    ):
        valid = (
            row[1] == "oracle_evaluations"
            and bool(str(row[2]).strip())
            and row[3] in {"fixed", "not_applicable", "unknown"}
            and ((row[3] == "fixed") == (row[4] is not None))
            and int(row[5]) > 0
            and _valid_json_container(row[6], list, non_empty=True)
            and _valid_json_container(row[7], dict)
            and row[8] is not None
        )
        if not valid:
            issues.append(str(row[0]))
    issues.extend(
        str(row[0])
        for row in connection.execute(
            """
            SELECT comparison.comparison_set_id
            FROM comparison_sets AS comparison
            LEFT JOIN comparison_set_members AS member USING (comparison_set_id)
            GROUP BY comparison.comparison_set_id
            HAVING COUNT(member.member_id) = 0
            """
        )
    )
    issues.extend(
        str(row[0])
        for row in connection.execute(
            """
            SELECT member.member_id
            FROM comparison_set_members AS member
            LEFT JOIN method_visualization_profiles AS profile
              ON member.method_id = profile.method_id AND member.profile_id = profile.profile_id
            WHERE profile.profile_id IS NULL
               OR trim(member.label) = ''
               OR member.display_order < 1
               OR NOT json_valid(member.parameters_json)
            """
        )
    )
    return issues


def _learning_edge_issues(connection: sqlite3.Connection) -> list[str]:
    tables = _table_names(connection)
    columns = {str(row[1]) for row in connection.execute("PRAGMA table_info(learning_edges)")}
    if "difficulty" not in columns:
        return _legacy_learning_edge_issues(connection)
    resolvers = {
        "method": ("methods", "method_id"),
        "problem": ("problem_archetypes", "problem_id"),
        "feature": ("problem_features", "feature_id"),
        "case": ("example_cases", "case_id"),
        "implementation": ("implementations", "implementation_id"),
        "view_preset": ("view_presets", "preset_id"),
        "scenario": ("demo_scenarios", "scenario_id"),
        "comparison": ("comparison_sets", "comparison_set_id"),
    }
    allowed_relations = {
        "prerequisite_for",
        "next_step",
        "contrast_with",
        "special_case_of",
        "generalizes",
        "applied_in",
        "common_misconception_for",
        "see_visualization",
        "see_comparison",
        "see_case",
        "implemented_by",
    }
    allowed_difficulties = {"beginner", "intermediate", "advanced", "all"}
    allowed_audiences = {"learner", "practitioner", "researcher", "all"}
    allowed_statuses = {"current", "deprecated", "draft"}
    issues: list[str] = []
    seen_semantics: set[tuple[str, str, str, str, str]] = set()
    for (
        edge_id,
        source_type,
        source_id,
        target_type,
        target_id,
        relation,
        rationale,
        difficulty,
        audience,
        display_order,
        source_ids_json,
        last_verified,
        status,
    ) in connection.execute(
        "SELECT edge_id, source_type, source_id, target_type, target_id, relation, "
        "rationale, difficulty, audience, display_order, source_ids_json, last_verified, status "
        "FROM learning_edges"
    ):
        edge = str(edge_id)
        semantics = (
            str(source_type),
            str(source_id),
            str(target_type),
            str(target_id),
            str(relation),
        )
        invalid_semantics = (
            str(relation) not in allowed_relations
            or not str(source_id).strip()
            or not str(target_id).strip()
            or (source_type == target_type and source_id == target_id)
            or not isinstance(rationale, str)
            or not rationale.strip()
            or difficulty not in allowed_difficulties
            or audience not in allowed_audiences
            or not isinstance(display_order, int)
            or isinstance(display_order, bool)
            or display_order < 1
            or not _valid_json_container(source_ids_json, list)
            or not json.loads(str(source_ids_json))
            or not str(last_verified).strip()
            or status not in allowed_statuses
            or semantics in seen_semantics
        )
        if invalid_semantics:
            issues.append(edge)
        seen_semantics.add(semantics)
        for endpoint_type, endpoint_id in ((source_type, source_id), (target_type, target_id)):
            resolver = resolvers.get(str(endpoint_type))
            if resolver is None:
                issues.append(edge)
                continue
            table, column = resolver
            if table not in tables:
                issues.append(edge)
                continue
            if (
                connection.execute(
                    f'SELECT 1 FROM "{table}" WHERE "{column}" = ?', (endpoint_id,)
                ).fetchone()
                is None
            ):
                issues.append(edge)
    return sorted(set(issues))


def _legacy_learning_edge_issues(connection: sqlite3.Connection) -> list[str]:
    """Validate the published 0.8 schema during the atomic 0.9 migration preflight."""

    resolvers = {
        "method": ("methods", "method_id"),
        "view_preset": ("view_presets", "preset_id"),
        "visualization_profile": ("method_visualization_profiles", "profile_id"),
        "objective": ("problem_instances", "problem_instance_id"),
        "scenario": ("demo_scenarios", "scenario_id"),
        "comparison": ("comparison_sets", "comparison_set_id"),
    }
    issues: list[str] = []
    for row in connection.execute(
        "SELECT edge_id, source_type, source_id, target_type, target_id, relation, rationale "
        "FROM learning_edges"
    ):
        if row[5] not in {"prerequisite", "next", "related", "contrast"} or not str(row[6]).strip():
            issues.append(str(row[0]))
        for endpoint_type, endpoint_id in ((row[1], row[2]), (row[3], row[4])):
            resolver = resolvers.get(str(endpoint_type))
            if (
                resolver is None
                or connection.execute(
                    f'SELECT 1 FROM "{resolver[0]}" WHERE "{resolver[1]}" = ?', (endpoint_id,)
                ).fetchone()
                is None
            ):
                issues.append(str(row[0]))
    return sorted(set(issues))


def _terminology_alias_issues(connection: sqlite3.Connection) -> list[str]:
    resolvers = {
        "method": ("methods", "method_id"),
        "problem": ("problem_archetypes", "problem_id"),
        "feature": ("problem_features", "feature_id"),
        "implementation": ("implementations", "implementation_id"),
    }
    issues: list[str] = []
    owners: dict[str, list[tuple[str, str, str | None]]] = defaultdict(list)
    json_columns = range(5, 10)
    for row in connection.execute(
        "SELECT term_id, target_type, target_id, label_ja, label_en, abbreviations_json, "
        "synonyms_json, domain_terms_json, misspellings_json, deprecated_terms_json, "
        "disambiguation_note, locale, rationale, source_ids_json, last_verified "
        "FROM terminology_aliases"
    ):
        term_id = str(row[0])
        resolver = resolvers.get(str(row[1]))
        containers_valid = all(_valid_json_container(row[index], list) for index in json_columns)
        source_ids = json.loads(str(row[13])) if _valid_json_container(row[13], list) else []
        target_exists = (
            resolver is not None
            and connection.execute(
                f'SELECT 1 FROM "{resolver[0]}" WHERE "{resolver[1]}" = ?', (row[2],)
            ).fetchone()
            is not None
        )
        sources_close = bool(source_ids) and all(
            connection.execute("SELECT 1 FROM sources WHERE source_id = ?", (source_id,)).fetchone()
            for source_id in source_ids
        )
        if (
            not target_exists
            or not str(row[3]).strip()
            or not str(row[4]).strip()
            or not containers_valid
            or not str(row[11]).strip()
            or not str(row[12]).strip()
            or not sources_close
            or not str(row[14]).strip()
        ):
            issues.append(term_id)
            continue
        terms = [str(row[3]), str(row[4])]
        for index in json_columns:
            terms.extend(str(value) for value in json.loads(str(row[index])))
        unique_terms = [value.casefold().strip() for value in terms]
        if len(unique_terms) != len(set(unique_terms)):
            issues.append(term_id)
        for term in (value.replace(" ", "") for value in unique_terms):
            owners[term].append((str(row[1]), str(row[2]), row[10]))
    for term, rows in owners.items():
        if len({(row[0], row[1]) for row in rows}) > 1 and any(row[2] is None for row in rows):
            issues.append(f"collision:{term}")
    return sorted(set(issues))


def _coverage_expectation_issues(connection: sqlite3.Connection) -> list[str]:
    resolvers = {
        "method": ("methods", "method_id"),
        "problem": ("problem_archetypes", "problem_id"),
        "feature_family": ("problem_features", "category"),
    }
    slices = {
        str(row[0]) for row in connection.execute("SELECT slice_id FROM learning_slice_priorities")
    }
    issues: list[str] = []
    for row in connection.execute(
        "SELECT expectation_id, subject_type, subject_id, applicability, rationale, "
        "source_ids_json, slice_id FROM learning_coverage_expectations"
    ):
        expectation_id = str(row[0])
        resolver = resolvers.get(str(row[1]))
        if resolver is None:
            issues.append(expectation_id)
            continue
        table, column = resolver
        subject_exists = connection.execute(
            f'SELECT 1 FROM "{table}" WHERE "{column}" = ? LIMIT 1', (row[2],)
        ).fetchone()
        source_ids = json.loads(str(row[5])) if _valid_json_container(row[5], list) else []
        sources_close = all(
            connection.execute("SELECT 1 FROM sources WHERE source_id = ?", (source_id,)).fetchone()
            for source_id in source_ids
        )
        if (
            subject_exists is None
            or row[3] not in {"expected", "not_applicable"}
            or not str(row[4]).strip()
            or not source_ids
            or not sources_close
            or (row[6] is not None and str(row[6]) not in slices)
        ):
            issues.append(expectation_id)
    for row in connection.execute(
        "SELECT slice_id, classification_score, misconception_score, visualization_score, "
        "demand_score, source_ids_json FROM learning_slice_priorities"
    ):
        scores = row[1:5]
        source_ids = json.loads(str(row[5])) if _valid_json_container(row[5], list) else []
        if (
            any(not isinstance(score, int) or not 0 <= score <= 3 for score in scores)
            or not source_ids
        ):
            issues.append(str(row[0]))
    return sorted(set(issues))


def _explicit_state_issues(
    connection: sqlite3.Connection, tables: set[str] | frozenset[str]
) -> list[str]:
    issues: list[str] = []
    state_columns = {
        "view_presets": {
            "root_support_status": {"supported", "unsupported", "unknown", "not_applicable"}
        },
        "method_visualization_profiles": {
            "support_status": {"supported", "unsupported", "unknown", "not_applicable"},
            "implementation_status": {
                "supported",
                "unsupported",
                "unknown",
                "not_applicable",
            },
        },
        "problem_definitions": {"objective_direction": {"minimize", "maximize", "multiobjective"}},
        "problem_instances": {
            "seed_status": {"fixed", "not_applicable", "unknown"},
            "known_reference_status": {
                "known_exact",
                "known_reference",
                "best_known",
                "unknown",
                "not_meaningful",
            },
        },
        "demo_scenarios": {"seed_status": {"fixed", "not_applicable", "unknown"}},
        "comparison_sets": {"seed_status": {"fixed", "not_applicable", "unknown"}},
    }
    for table, columns in state_columns.items():
        if table not in tables:
            continue
        for column, allowed in columns.items():
            placeholders = ", ".join("?" for _ in allowed)
            count = int(
                connection.execute(
                    f'SELECT COUNT(*) FROM "{table}" '
                    f'WHERE "{column}" IS NULL OR trim("{column}") = \'\' '
                    f'OR "{column}" NOT IN ({placeholders})',
                    tuple(sorted(allowed)),
                ).fetchone()[0]
            )
            if count:
                issues.append(f"{table}.{column}:{count}")
    return issues


def _versioned_claim_issues(connection: sqlite3.Connection, tables: list[str]) -> list[str]:
    if "implementation_claims" not in tables:
        return ["missing:implementation_claims"]
    issues: list[str] = []
    implementation_count = int(
        connection.execute("SELECT COUNT(*) FROM implementations").fetchone()[0]
    )
    active_count = int(
        connection.execute(
            "SELECT COUNT(*) FROM implementation_claims WHERE valid_to IS NULL"
        ).fetchone()[0]
    )
    expected = implementation_count * 7
    if active_count != expected:
        issues.append(f"active-claim-count:{active_count}/{expected}")
    duplicates = connection.execute(
        """
        SELECT subject_id, predicate FROM implementation_claims
        WHERE valid_to IS NULL GROUP BY subject_id, predicate HAVING COUNT(*) <> 1
        """
    ).fetchall()
    issues.extend(f"active-duplicate:{row[0]}:{row[1]}" for row in duplicates)
    broken_replacements = int(
        connection.execute(
            """
            SELECT COUNT(*) FROM implementation_claims AS old
            LEFT JOIN implementation_claims AS new ON new.claim_id = old.replaced_by
            WHERE old.verification_status = 'superseded'
              AND (new.claim_id IS NULL OR old.valid_to IS NULL OR new.valid_from <= old.valid_to)
            """
        ).fetchone()[0]
    )
    if broken_replacements:
        issues.append(f"broken-supersession:{broken_replacements}")
    placeholders = ",".join("?" for _ in HIGH_USAGE_IMPLEMENTATION_IDS)
    high_usage_total = len(HIGH_USAGE_IMPLEMENTATION_IDS)
    high_usage_covered = int(
        connection.execute(
            f"""
            SELECT COUNT(*) FROM implementation_claims
            WHERE subject_id IN ({placeholders}) AND predicate = 'current_release'
              AND valid_to IS NULL AND value_status = 'verified'
            """,
            tuple(sorted(HIGH_USAGE_IMPLEMENTATION_IDS)),
        ).fetchone()[0]
    )
    if high_usage_total and (100 * high_usage_covered / high_usage_total) < 80:
        issues.append(f"high-usage-release-coverage:{high_usage_covered}/{high_usage_total}")
    return issues


def _benchmark_context_issues(connection: sqlite3.Connection, tables: list[str]) -> list[str]:
    if "benchmark_contexts" not in tables:
        return ["missing:benchmark_contexts"]
    issues: list[str] = []
    categories = {
        str(row[0]) for row in connection.execute("SELECT category FROM benchmark_contexts")
    }
    missing = {"LP", "QP", "NLP", "MIP", "DFO", "BO"} - categories
    issues.extend(f"missing-category:{category}" for category in sorted(missing))
    contextless = connection.execute(
        "SELECT comparison_set_id FROM comparison_sets WHERE benchmark_context_id IS NULL"
    ).fetchall()
    issues.extend(f"contextless-comparison:{row[0]}" for row in contextless)
    blank_json = connection.execute(
        """
        SELECT context_id FROM benchmark_contexts
        WHERE json_array_length(source_ids_json) = 0
           OR json_array_length(outcome_metrics_json) = 0
           OR status_mapping_json = '{}'
           OR implementation_versions_json = '{}'
        """
    ).fetchall()
    issues.extend(f"incomplete-context:{row[0]}" for row in blank_json)
    return issues


def _structured_failure_mode_issues(connection: sqlite3.Connection, tables: list[str]) -> list[str]:
    required = {
        "failure_mode_profiles",
        "failure_mode_triggers",
        "failure_mode_symptoms",
        "failure_mode_diagnostics",
        "failure_mode_mitigations",
        "failure_mode_affected_entities",
        "failure_mode_scenarios",
    }
    missing = required - set(tables)
    if missing:
        return [f"missing:{table}" for table in sorted(missing)]
    issues: list[str] = []
    profile_count = int(
        connection.execute("SELECT COUNT(*) FROM failure_mode_profiles").fetchone()[0]
    )
    if profile_count < 12:
        issues.append(f"profile-count:{profile_count}/12")
    for table in (
        "failure_mode_triggers",
        "failure_mode_symptoms",
        "failure_mode_diagnostics",
        "failure_mode_mitigations",
        "failure_mode_affected_entities",
    ):
        absent = connection.execute(
            f"""SELECT profile.failure_mode_id FROM failure_mode_profiles AS profile
                LEFT JOIN {table} AS child USING (failure_mode_id)
                GROUP BY profile.failure_mode_id HAVING COUNT(child.failure_mode_id) = 0"""
        ).fetchall()
        issues.extend(f"missing-{table}:{row[0]}" for row in absent)
    unresolved_methods = connection.execute(
        """SELECT affected.failure_mode_id, affected.entity_id
           FROM failure_mode_affected_entities AS affected
           LEFT JOIN methods ON affected.entity_type = 'method'
             AND methods.method_id = affected.entity_id
           WHERE affected.entity_type = 'method' AND methods.method_id IS NULL"""
    ).fetchall()
    issues.extend(f"missing-method:{row[0]}:{row[1]}" for row in unresolved_methods)
    unresolved_implementations = connection.execute(
        """SELECT affected.failure_mode_id, affected.entity_id
           FROM failure_mode_affected_entities AS affected
           LEFT JOIN implementations ON affected.entity_type = 'implementation'
             AND implementations.implementation_id = affected.entity_id
           WHERE affected.entity_type = 'implementation'
             AND implementations.implementation_id IS NULL"""
    ).fetchall()
    issues.extend(f"missing-implementation:{row[0]}:{row[1]}" for row in unresolved_implementations)
    unresolved_features = connection.execute(
        """SELECT affected.failure_mode_id, affected.entity_id
           FROM failure_mode_affected_entities AS affected
           LEFT JOIN problem_features ON affected.entity_type = 'feature'
             AND problem_features.feature_id = affected.entity_id
           WHERE affected.entity_type = 'feature' AND problem_features.feature_id IS NULL"""
    ).fetchall()
    issues.extend(f"missing-feature:{row[0]}:{row[1]}" for row in unresolved_features)
    unresolved_sources = connection.execute(
        """SELECT profile.failure_mode_id, source.value
           FROM failure_mode_profiles AS profile, json_each(profile.source_ids_json) AS source
           LEFT JOIN sources ON sources.source_id = source.value
           WHERE sources.source_id IS NULL"""
    ).fetchall()
    issues.extend(f"missing-source:{row[0]}:{row[1]}" for row in unresolved_sources)
    observable_rows = connection.execute(
        "SELECT failure_mode_id, observable_id FROM failure_mode_symptoms "
        "WHERE observable_id IS NOT NULL"
    ).fetchall()
    issues.extend(
        f"unknown-observable:{row[0]}:{row[1]}"
        for row in observable_rows
        if str(row[1]) not in OBSERVABLE_IDS
    )
    scenario_rows = connection.execute(
        "SELECT failure_mode_id, scenario_id FROM failure_mode_scenarios"
    ).fetchall()
    if len(scenario_rows) < 4:
        issues.append(f"failure-scenario-count:{len(scenario_rows)}/4")
    expected_scenarios = set(FAILURE_SCENARIOS.values())
    issues.extend(
        f"unknown-scenario:{row[0]}:{row[1]}"
        for row in scenario_rows
        if str(row[1]) not in expected_scenarios
    )
    misclassified = connection.execute(
        """SELECT affected.failure_mode_id FROM failure_mode_affected_entities AS affected
           JOIN failure_mode_profiles AS profile USING (failure_mode_id)
           WHERE profile.failure_scope = 'implementation_specific'
             AND affected.specificity <> 'implementation_only'"""
    ).fetchall()
    issues.extend(f"implementation-scope-leak:{row[0]}" for row in misclassified)
    return issues
